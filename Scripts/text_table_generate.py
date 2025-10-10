import argparse
from enum import IntEnum
from openpyxl import Workbook, load_workbook
from pathlib import Path


# Формирование строки для одного элемента(сигнала/агрегата) раздела таблицы
def concate_name(poz, info, type=0, text="", separator=", "):
    match type:
        case 0:
            if info and poz and info.lower() != "none" and poz.lower() != "none":
                name_str = poz + separator + info
            elif poz and poz.lower() != "none":
                name_str = poz
            else:
                name_str = info
            return name_str
        case 1:
            if poz and poz.lower() != "none":
                name_str = poz
            else:
                name_str = info
            return name_str
        case 2:
            if info and info.lower() != "none":
                name_str = info
            else:
                name_str = poz
            return name_str
        case 4:
            return text


# Функция создания одного раздела
def create_order(wb_data, ws_name, name, last_name, name_num,
                 name_order=list(range(0, 50)), start_id=0,
                 type=0, text="", separator=", ", count_for_type4=16):
    ws_data = wb_data[name]
    id = start_id

    for row in range(start_id):
        if row == 0:
            ws_name.append([name_order[name_num], name + last_name, row])
        else:
            ws_name.append([name_order[name_num], "", row])

    for row in list(ws_data.rows)[1:]:  # Начинаем со второй строки таблицы
        poz = str(row[0].value)  # Позиционное обозначение
        info = str(row[1].value)  # Имя сигнала

        name_str = concate_name(poz, info, type, text, separator)

        if id == 0:
            ws_name.append([name_order[name_num], name + last_name, id, name_str])
            id += 1

        if type < 4:
            count = 1
            suffixes = [""]
        else:
            count = count_for_type4
            suffixes = map(str, range(1, count + 1))

        for suffix in suffixes:
            ws_name.append([name_order[name_num], "", id, name_str + " " + suffix])
            id += 1


def create_name_table(out_map_path, data_path, file_name="name_table.xlsx",
                      name_order=list(range(0, 20)), start_id=0, enable_export=False):

    class NameType(IntEnum):
        FULL = 0
        TAG = 1
        INFO = 2
        TYPE4 = 4

    if not enable_export:
        return

    wb_data = load_workbook(data_path)
    wb_name = Workbook()
    ws_name = wb_name.active
    ws_name.append([
        "ID раздела", "Описание", "ID строки", "Language 1", "Language 2", "Language 3",
        "Language 4", "Language 5", "Language 6", "Language 7", "Language 8"
    ])

    name_num = 0
    list_all_ws = wb_data.sheetnames

    # --- Функция для обработки листа ---
    def process_sheet(name, suffix, type_, text=""):
        nonlocal name_num
        if name not in list_all_ws:
            print(f"Лист {name} не найден в файле данных.")
            return
        create_order(wb_data, ws_name, name, suffix, name_num, name_order, start_id, type=type_, text=text)
        name_num += 1

    # --- Сигналы (полные названия) ---
    for name in ["AI", "AO", "DI", "DO"]:
        process_sheet(name, "", NameType.FULL)

    # --- Сигналы (тэг + описание отдельно) ---
    for name in ["AI", "AO", "DI", "DO"]:
        process_sheet(name, "_TAG", NameType.TAG)
        process_sheet(name, "_INFO", NameType.INFO)

    # --- Агрегаты ---
    for name in ["DVLV", "AVLV", "MTR", "MTRPID"]:
        process_sheet(name, "_TAG", NameType.TAG)
        process_sheet(name, "_INFO", NameType.INFO)

    # --- Блокировки ---
    process_sheet("BL", "_TAG", NameType.FULL)
    process_sheet("BL", "_INFO", NameType.TYPE4, text="Название блокировки")

    # Сохраняем
    output_path = Path(out_map_path) / file_name
    wb_name.save(output_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate name table from signals Excel file.")
    parser.add_argument("out_map_path", type=str, help="Output folder path")
    parser.add_argument("data_path", type=str, help="Input Excel file path (e.g., signals.xlsx)")
    parser.add_argument("--file_name", type=str, default="name_table.xlsx", help="Output file name")
    parser.add_argument("--start_id", type=int, default=0, help="Starting ID for rows")
    parser.add_argument("--enable_export", action="store_true", help="Enable export (default: True)")

    args = parser.parse_args()

    create_name_table(
        out_map_path=args.out_map_path,
        data_path=args.data_path,
        file_name=args.file_name,
        name_order=list(range(0, 30)),
        start_id=args.start_id,
        enable_export=args.enable_export
    )