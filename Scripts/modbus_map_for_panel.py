import openpyxl
import re

def convert_modbus_map(path_map="",name_new_map="modbus_for_panel", name_sheet = "Шаблон 2"):
    try:
        workbook = openpyxl.load_workbook(path_map)
    except Exception as e:
        print(f"Ошибка при открытии файла: {e}")
        return False

    if name_sheet not in workbook.sheetnames:
        print(f"Лист '{name_sheet}' не найден в таблице")
        return False

    ws_data = workbook[name_sheet]

    # Именованные столбцы для ясности
    ADDR_LABEL = 'K'  # Адресная метка
    DEVICE_NAME = 'J'  # Имя устройства
    BITMAP_TYPE = 'L'  # Тип: BITMAP или нет
    DATA_TYPE = 'D'  # Тип данных (BOOL и др.)
    REG_ADDR = 'B'  # Адрес регистра
    SIGNAL_NAME = 'A'  # Имя сигнала
    CATEGORY = 'N' # Категория
    PRIORITY = 'O' # Приоритет
    CONDITION = 'P' # Условие
    EXPLAN    = 'Q' # Пояснение регистра (Для битотвых масок)


    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Sheet 1"

    name_plc = "PLC"
    row_counter = 1

    def get_function_type(data_type: str):
        return "4x_Bit" if data_type.lower() == "bool" else "4x"

    def convert_data_type(data_type: str) -> str:
        if not data_type:
            return "Неопределенный"

        data_type = data_type.strip()

        mapping = {
            "FLOAT(4 byte)": "32-bit Float",
            "WORD(2 byte)": "16-bit Unsigned",
            "UDINT(4 byte)": "32-bit Unsigned",
            "4 byte": "32-bit Unsigned",
            "2 byte": "16-bit Unsigned"
        }

        # 1. Точное совпадение
        if data_type in mapping:
            return mapping[data_type]

        # 2. Частичное совпадение — с приоритетом длинных ключей
        for key in sorted(mapping.keys(), key=len, reverse=True):
            if key in data_type:
                return mapping[key]

        return "Неопределенный"

    max_row = ws_data.max_row
    for row in range(2, max_row + 1):
        addr_label_val = ws_data[f"{ADDR_LABEL}{row}"].value
        if not addr_label_val:  # Пропускаем пустые строки
            continue

        bitmap_type_val = ws_data[f"{BITMAP_TYPE}{row}"].value
        device_name_val = ws_data[f"{DEVICE_NAME}{row}"].value or ""
        reg_addr_val = ws_data[f"{REG_ADDR}{row}"].value
        data_type_val = ws_data[f"{DATA_TYPE}{row}"].value
        panel_data_type = convert_data_type(data_type_val)
        signal_name_val = ws_data[f"{SIGNAL_NAME}{row}"].value
        category_name_val = ws_data[f"{CATEGORY}{row}"].value
        priority_name_val = ws_data[f"{PRIORITY}{row}"].value
        condition_name_val = ws_data[f"{CONDITION}{row}"].value
        explan_val = ws_data[f"{EXPLAN}{row}"].value

        # Определяем тип функции
        func_type = get_function_type(data_type_val)

        if bitmap_type_val == "BITMAP":
            # Обработка BITMAP: разбиваем по строкам
            try:
                def cell_bitmap(val: str) -> dict:  # Изменим возвращаемый тип на dict, так как это словарь
                    result_dict = {}  # Создаем локальную переменную
                    if val is not None:  # Добавим проверку на None
                        for line in str(val).strip().split(
                                '\n'):  # str() не обязателен, если val точно str, но не мешает
                            if '-' in line:
                                try:  # Добавим обработку ошибок на случай некорректного формата строки
                                    bit, name_part = line.split('-', 1)
                                    result_dict[int(bit.strip())] = name_part.strip()  # Преобразуем бит в int
                                except ValueError:
                                    # Можно вывести предупреждение или просто игнорировать строку
                                    # print(f"Предупреждение: Невозможно разобрать строку '{line}'")
                                    continue
                    return result_dict  # Возвращаем созданный словарь

                # Сортируем по биту
                label_list = sorted(cell_bitmap(addr_label_val).items())
                priority_list = sorted(cell_bitmap(priority_name_val).items())
                explan_list = sorted(cell_bitmap(explan_val).items())

                #print(label_list, priority_list, explan_list)

                #for bit, bit_name in label_list:
                for i in range(len(label_list)):

                    bit = label_list[i][0]
                    bit_name  = label_list[i][1]
                    priority = priority_list[i][1] if priority_list else ""
                    explan = explan_list[i][1] if explan_list else ""

                    def parse_mode_string(s):
                        # Разбиваем по точке с запятой
                        pattern = r'([^;]+?)\s*\(\s*(\d+)\s*\)'
                        matches = re.findall(pattern, s)
                        result = {}
                        for label, digit in matches:
                            result[int(digit)] = label.strip()

                        if result:
                            return result
                        return ""

                    explan_state = parse_mode_string(explan)

                    signal_full_name = f"{bit_name}_{device_name_val}" if device_name_val else bit_name

                    new_sheet.cell(row=row_counter, column=1, value=signal_full_name)  # A: имя сигнала
                    new_sheet.cell(row=row_counter, column=2, value=name_plc)  # B: имя устройства
                    new_sheet.cell(row=row_counter, column=3, value="4x_Bit")  # C: тип (всегда 4x_Bit для BITMAP)
                    new_sheet.cell(row=row_counter, column=4, value=f"{reg_addr_val}.{bit}")  # D: адрес
                    new_sheet.cell(row=row_counter, column=6, value=panel_data_type)
                    new_sheet.cell(row=row_counter, column=7, value=category_name_val)
                    new_sheet.cell(row=row_counter, column=8, value=priority)
                    new_sheet.cell(row=row_counter, column=9, value=condition_name_val)

                    if explan_state:
                        print(explan_state)
                        new_sheet.cell(row=row_counter, column=5, value=f"{signal_name_val}.{explan_state}")
                    else:
                        new_sheet.cell(row=row_counter, column=5, value=signal_name_val)

                    row_counter += 1
            except Exception as e:
                print(f"Ошибка обработки BITMAP в строке {row}: {e} .")
                continue
        else:
            # Обычная строка
            signal_full_name = f"{addr_label_val}_{device_name_val}" if device_name_val else addr_label_val

            new_sheet.cell(row=row_counter, column=1, value=signal_full_name)
            new_sheet.cell(row=row_counter, column=2, value=name_plc)
            new_sheet.cell(row=row_counter, column=3, value=func_type)
            new_sheet.cell(row=row_counter, column=4, value=reg_addr_val)
            new_sheet.cell(row=row_counter, column=5, value=signal_name_val)
            new_sheet.cell(row=row_counter, column=6, value=panel_data_type)
            new_sheet.cell(row=row_counter, column=7, value=category_name_val)
            new_sheet.cell(row=row_counter, column=8, value=priority_name_val)
            new_sheet.cell(row=row_counter, column=9, value=condition_name_val)

            row_counter += 1

    # Сохранение
    try:
        new_workbook.save(name_new_map + ".xlsx")  # или просто name_new_map, если уже с .xlsx
        print(f"Файл успешно сохранён как {name_new_map}.xlsx")
        return True
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")
        return False



if __name__ == "__main__" :
    convert_modbus_map("C:/Users/deminid/PycharmProjects/ModbusApp/Example file/modbus_map_test.xlsx", name_sheet="Шаблон")