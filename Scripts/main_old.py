import pdfplumber as pl
import re
from collections import namedtuple
import pandas as pd
import openpyxl
import numpy
from copy import copy
import dearpygui.dearpygui as dpg


def pdf_to_xlsx(path_pdf, path_xlsx, start_page, end_page):
    di_map = namedtuple("di_map", "poz name module_name module_poz channel contact type_signal".split())
    ai_map = namedtuple("ai_map", "poz name module_name module_poz channel contact type_signal".split())
    do_map = namedtuple("do_map", "poz name module_name module_poz channel contact type_signal".split())
    ao_map = namedtuple("ao_map", "poz name module_name module_poz channel contact type_signal".split())
    maps = [di_map, ai_map, do_map, ao_map]  # список кортежей
    names = ["DI", "AI", "DO", "AO"]  # список имен
    tables = [list() for x in range(len(maps))]  # список таблиц

    im_map = namedtuple("im_map", "poz name".split())
    im_names = ["DVLV", "AVLV", "MTR", "MTRPID", "BL"]
    im_tables = [list() for x in range(len(im_names))]
    im_tables[0].append(im_map("Поз1", "Задвижка 1"))
    im_tables[1].append(im_map("Поз1", "Регулируемый клапан 1"))
    im_tables[2].append(im_map("Поз1", "Агрегат 1"))
    im_tables[3].append(im_map("Поз1", "Регулируемый мотор 1"))
    im_tables[4].append(im_map("Поз1", "Блокировка 1"))

    with pl.open(path_pdf) as pdf:  # Открытие pdf файла конструкторской документации
        pages = pdf.pages  # Извлекаем страницы pdf
        for page in pages[(start_page - 1):end_page]:  # Проходимся по каждой странице
            table = page.extract_table()  # Извлекаем таблицу на странице
            for line in table:  # Проходимся по строкам в таблице
                line = [x for x in line if x is not None]  # Удаляем все None из строки
                for i in range(4):  # Проверяем строку 4 раза
                    if re.search(f"{names[i]}" + r"\d{1,2}",
                                 str(line)):  # Есть ли в строке обозначение канала из списка (["DI", "AI", "DO", "AO"])
                        if re.search(r"№", line[0]) and len(line) > 12:  # удаление подписи из строки (если попалась №)
                            del line[:2]  # удаляем два первых элемента

                        poz = line[0]  # извлекаем позиционное обозначение
                        name_param = line[1].replace("\n", " ")  # название параметра
                        module_name = line[-1].replace("\n", " ")  # название модуля
                        module_poz = line[-3]  # позиционное обозначение модуля
                        channel = line[-2]  # наименование канала

                        # Перемечки распознаются как две ячейки , поэтому длина массива увеличивается на 1
                        if len(line) == 18 :
                            type = line[-9]  # тип сигнала
                            contact = line[-7]  # контакт
                        else :
                            type = line[-8]  # тип сигнала
                            contact = line[-6]  # контакт

                        if re.search(r"резерв", poz, flags=re.IGNORECASE) or re.search(r"резерв", name_param,
                                                                                       flags=re.IGNORECASE):
                            poz = "РЕЗЕРВ"
                            name_param = " "
                        tables[i].append(maps[i](poz, name_param, module_name, module_poz, channel, contact,
                                                 type))  # Добавляем полученные данные в виде кортежа в соответвующий список

        with pd.ExcelWriter(path_xlsx) as writer:  # Открытие и изменение xlsx файла (либо создание если нет)
            for i in range(len(maps)):  # Проходимся 4 раза
                df = pd.DataFrame(tables[i])  # Преобразуем список в датафрейм
                print(df)  #
                df.to_excel(writer, sheet_name=names[i], index=False)  # Записываем таблицу на соответвующий лист
            for i in range(len(im_tables)):
                df = pd.DataFrame(im_tables[i])
                df.to_excel(writer, sheet_name=im_names[i], index=False)  # Записываем таблицу на соответвующий лист
    return True


def data_find(data_path):
    IO_num = {}

    wb_data = openpyxl.load_workbook(data_path)
    for sheet in wb_data.sheetnames:
        ws_data = wb_data[sheet]
        count = ws_data.max_row - 2  # Вычитаем строку наименования и единицу , так как в шаблоне уже есть одна строка
        IO_num.update([(sheet, count)])
    # print(IO_num)
    return IO_num


import openpyxl

def marker_find(sample_path, name_sheet):
    markers = []
    size = []
    start_adress = []
    number_cell_marker = -1
    number_row = 0

    # Загружаем с data_only=True, чтобы получать результаты формул, а не сами формулы
    wb_sample = openpyxl.load_workbook(sample_path, data_only=True)
    ws_sample = wb_sample[name_sheet]

    for row in ws_sample.rows:
        if number_cell_marker == -1:
            for cell in row:
                if cell.value == "МЕТКА":
                    number_cell_marker = cell.column
                    break  # выходим из цикла, как только нашли
        else:
            marker_cell = row[number_cell_marker - 1]
            marker = marker_cell.value
            if marker is not None:  # проверяем, что значение не None
                markers.append(marker)
                # row[1].row — это номер текущей строки (начиная с 1)
                current_row_num = row[1].row
                size.append(current_row_num - number_row - 1)
                number_row = current_row_num
                # Получаем значение из следующей ячейки (предположительно адрес)
                if number_cell_marker < len(row):
                    start_adress.append(row[number_cell_marker].value)
                else:
                    start_adress.append(None)

    # Убираем последний маркер и первый размер (как в оригинале)
    if markers:
        markers = markers[:-1]
    if size:
        size = size[1:]

    print(markers, size, start_adress)
    return markers, size, start_adress

def create_modbus_map(sample_map_path, out_map_path, name_sheet,
                      data_path, show_progress=False, progress_bar_tag="",
                      delete_other_sheets = False):
    # print(sample_map_path, out_map_path, name_sheet, data_path)
    size_val_modbus = 1
    bit_adress_modbus = 0
    bit_modbus = 0

    IO_num = data_find(data_path)
    markers, size, start_modbus_adress = marker_find(sample_map_path, name_sheet)
    # Находим маркеры в шаблоне, размер одного блока и начальный адрес
    num_channel = [IO_num.get(marker.split("_")[0]) for marker in markers if IO_num.get(marker.split("_")[0]) != None]
    # Формируем список количества каналов для каждого типа канала по списку маркеров

    wb_data = openpyxl.load_workbook(data_path)  # Открываем книгу с данными

    wb_sample = openpyxl.load_workbook(sample_map_path)  # Открываем шаблоны модбас карты
    list_all_ws = wb_sample.sheetnames  # Считываем список шаблонов
    # print(list_all_ws)

    if delete_other_sheets :
        for item in list_all_ws:  #
            if item != name_sheet:  #
                wb_sample.remove(wb_sample[item])  # Удаляем все листы кроме листа с нужным шаблоном

    ws_sample = wb_sample[str(name_sheet)]  #
    print(markers)
    k = 0
    for marker in markers:
        print(marker)
        type_channel = marker.split("_")[0]
        list_all_ws = wb_data.sheetnames
        if type_channel not in list_all_ws:  # Если маркера нет в списке листов таблицы данных
            print(type_channel)
            continue
        ws_data = wb_data[type_channel]
        row_num_data = {"AI": 2, "AO": 2, "DI": 2, "DO": 2, "DVLV": 2, "AVLV": 2, "MTR": 2,
                        "BL": 2, "MTRPID": 2}  # Номер первой строки в таблице исходных данных

        bit_adress_modbus = 0
        row_num = 0
        for row in ws_sample.rows:
            row_num += 1
            print(row_num)
            if row[8].value == marker:  # Находим строчку с маркером

                # Добавляем новые пустые строки
                ws_sample.insert_rows(row_num + 1 + size[k], num_channel[k] * size[k])

                for n in range(num_channel[k] + 1):  # Проходимся по кол-ву каналов одного типа строки

                    for j in range(size[k] * n + row_num + 1,
                                   size[k] * (n + 1) + row_num + 1):  # Проходимся по строчкам описания одного канала

                        # Указываем в ячейке 10(J) порядковый номер канала/задвижки/агрегата (нумерация с 1)
                        ws_sample.cell(row=j, column=10).value = f"{type_channel}{n+1}"

                        # Копируем название обозначения в 11(K) ячейку
                        if n != num_channel[k]:
                            old_cell = ws_sample.cell(row=j, column=11)  # Копируем старую ячейку
                            ws_sample.cell(row=j + size[k],
                                           column=11).value = old_cell.value  # Копируем значение ячекий
                            ws_sample.cell(row=j + size[k], column=11)._style = copy(
                                old_cell._style)  # Копируем стиль ячейки

                        # Копируем тип обработки метки (BITMAP) в 12(L) ячейку
                            old_cell = ws_sample.cell(row=j, column=12)  # Копируем старую ячейку
                            ws_sample.cell(row=j + size[k],
                                           column=12).value = old_cell.value  # Копируем значение ячекий
                            ws_sample.cell(row=j + size[k], column=12)._style = copy(
                                old_cell._style)  # Копируем стиль ячейки

                        for i in range(1, 8 + 1):  # Проходимся по ячейкам в строке
                            if n != num_channel[k]:
                                old_cell = ws_sample.cell(row=j, column=i)  # Копируем старую ячейку
                                ws_sample.cell(row=j + size[k],
                                               column=i).value = old_cell.value  # Копируем значение ячекий
                                ws_sample.cell(row=j + size[k], column=i)._style = copy(
                                    old_cell._style)  # Копируем стиль ячейки

                            if i == 1:  # Столбец 1(Название)
                                old_value = str(ws_sample.cell(row=j, column=i).value)
                                new_value = (str(ws_data.cell(row=row_num_data[type_channel], column=1).value) +
                                             "-" + str(ws_data.cell(row=row_num_data[type_channel], column=2).value))
                                new_value = old_value.replace("$$", new_value)
                                ws_sample.cell(row=j, column=i).value = new_value
                            if i == 2:  # Столбец 2(Адресс)
                                ws_sample.cell(row=j,
                                               column=2).value = f"{start_modbus_adress[k]}.{bit_adress_modbus}"
                                old_cell_data_type = str(ws_sample.cell(row=j, column=4).value)
                                new_cell_data_type = str(ws_sample.cell(row=j + 1, column=4).value)
                                if old_cell_data_type.find("4 byte") != -1:  # предыдущая строчка - 2 регистра
                                    size_val_modbus = 2
                                    bit_modbus = 0
                                    if new_cell_data_type.find("BOOL") != -1:  # новая строчка - булевское
                                        start_modbus_adress[k] += size_val_modbus
                                    ws_sample.cell(row=j, column=2).value = f"{start_modbus_adress[k]}"
                                elif old_cell_data_type.find("2 byte") != -1:  # предыдущая строчка - 1 регистр
                                    size_val_modbus = 1
                                    bit_modbus = 0
                                    if new_cell_data_type.find("BOOL") != -1:  # новая строчка - булевское
                                        start_modbus_adress[k] += size_val_modbus
                                    ws_sample.cell(row=j, column=2).value = f"{start_modbus_adress[k]}"
                                elif old_cell_data_type.find("BOOL") != -1:  # предыдущая строчка - булевское
                                    size_val_modbus = 0
                                    bit_modbus = 1
                                    if new_cell_data_type.find("4 byte") != -1:  # новая строчка - 2 регистра
                                        size_val_modbus = -1
                                        ws_sample.cell(row=j,
                                                       column=2).value = f"{start_modbus_adress[k]}.{bit_adress_modbus}"
                                else:
                                    ws_sample.cell(row=j,
                                                   column=2).value = f"{start_modbus_adress[k]}.{bit_adress_modbus}"
                            if i == 6:  # 6 столбец (Номер канала)
                                channel_name = (f"{ws_data.cell(row=row_num_data[type_channel], column=5).value}" +
                                                f"({ws_data.cell(row=row_num_data[type_channel], column=4).value})")
                                ws_sample.cell(row=j, column=i).value = channel_name

                        start_modbus_adress[k] += size_val_modbus
                        bit_adress_modbus += bit_modbus

                        if ((size[k] > 1 and bit_adress_modbus >= (4 * size[k]))                #Если
                                or bit_adress_modbus > 15                                       #
                                or (type_channel == "DVLV" and bit_adress_modbus >= size[k])    #
                                or (type_channel == "AI" and bit_adress_modbus >= size[k])  #
                                or (16 // size[k] == 1 and bit_adress_modbus >= size[k])):        #
                            start_modbus_adress[k] += 1
                            bit_adress_modbus = 0

                    row_num_data[type_channel] += 1

        k += 1
        progress = round(k / len(markers) * 100, 1)
        if show_progress:
            dpg.set_value(progress_bar_tag, k / len(markers))
            dpg.configure_item(progress_bar_tag, show=True, overlay=f"{progress}%")
            #print(marker)
        else:
            print(progress, "%")


    wb_sample.save(out_map_path)  # Сохраняем карту как новый файл
    return True


def create_name_table(out_map_path, data_path, file_name="name_table.xlsx",
                      name_order = [0, 1, 2, 3] , start_id = 0, enable_export=False):
    names = ["AI", "AO", "DI", "DO"]  # список имен
    order = name_order
    if enable_export:
        wb_data = openpyxl.load_workbook(data_path)  # Открываем книгу с данными
        wb_name = openpyxl.Workbook()  # Открываем книгу с данными

        ws_name = wb_name.worksheets[0]
        ws_name.append(["ID раздела", "Описание", "ID строки", "Language 1", "Language 2", "Language 3",
                        "Language 4", "Language 5", "Language 6", "Language 7", "Language 8"])
        name_num = 0  # Номер имени в списке
        for name in names:
            list_all_ws = wb_data.sheetnames
            if name not in list_all_ws:  # Если маркера нет в списке листов таблицы данных
                print(name)
                continue
            ws_data = wb_data[name]
            row_count = ws_data.max_row - 1  # кол-во сигналов одного типа
            row_num = start_id
            for row in list(ws_data.rows)[1:]:  # Начинаем со второй строки таблицы
                poz = str(row[0].value)  # Позиционное обозначение
                name_sig = str(row[1].value)  # Имя сигнала

                if name_sig and poz and name_sig.lower() != "none" and poz.lower() != "none":
                    name_str = poz + ", " + name_sig
                elif poz and poz.lower() != "none":
                    name_str = poz
                else:
                    name_str = name_sig

                if row_num == 0:
                    ws_name.append([order[name_num], name, row_num, name_str])
                else:
                    ws_name.append([order[name_num], "", row_num, name_str])
                row_num += 1
            name_num += 1
        wb_name.save(out_map_path + "\\" + file_name)


def gui():
    global path_data
    global path_sample
    global path_export
    global paths
    paths = ["", "", ""]
    global sample_list
    global sample_name
    global data_numbers
    data_numbers = {"start": 1, "end": 1}
    global error_str

    window_width = 800
    window_height = 800

    def create_map(sender, appdata):
        dpg.configure_item("load_indic_2", show=True)
        try:
            create_modbus_map(paths[1], paths[2] + "\\modbus_map.xlsx", sample_name, paths[2] + "\\signals.xlsx",
                              show_progress=True, progress_bar_tag="progress_bar_map")
        except Exception as ex:
            print(Exception)
            error_str = str(ex)
            dpg.set_value("error_text", error_str)
        dpg.configure_item("load_indic_2", show=False)

    def create_xlsx(sender, app_data):
        dpg.configure_item("load_indic_1", show=True)
        try:
            pdf_to_xlsx(paths[0], paths[2] + "\\signals.xlsx", data_numbers["start"], data_numbers["end"])
        except Exception as ex:
            print(Exception)
            error_str = str(ex)
            dpg.set_value("error_text", error_str)
        dpg.configure_item("load_indic_1", show=False)

    def create_table(sender, appdata):
        create_name_table(paths[2], paths[2] + "\\modbus_map.xlsx", "../name_table.xlsx", enable_export=True)

    def set_numbers(sender, app_data):
        data_numbers[sender] = app_data
        print(data_numbers[sender])

    def set_sample_name(sender, app_data):
        global sample_name
        sample_name = str(app_data)
        print(sample_name)

    def sample_list_extr(sender, app_data):
        try:
            path_sample = paths[1]
            wb_sample = openpyxl.load_workbook(path_sample)
            global sample_list
            sample_list = wb_sample.sheetnames
            dpg.configure_item("sample_combo", items=sample_list)
        except Exception as ex:
            error_str = str(ex)
            dpg.set_value("error_text", error_str)

    def get_pdf_page_count(pdf_path):
        """Возвращает количество страниц в PDF файле с использованием pdfplumber"""
        try:
            with pl.open(pdf_path) as pdf:
                return len(pdf.pages)
        except Exception as ex:
            print(f"Ошибка при чтении PDF: {ex}")
            return 1  # Возвращаем 1 в случае ошибки, чтобы интерфейс не сломался

    def path_extractor(sender, app_data, id):
        try:
            if sender == "filedialog_2":
                paths[id] = app_data["current_path"]
            else:
                paths[id] = str(list(app_data["selections"].values())[0])
            dpg.set_value(f"input_text_{id}", paths[id])

            # Если выбран PDF (id=0), обновляем максимальное количество страниц
            if id == 0 and paths[0].lower().endswith('.pdf'):
                total_pages = get_pdf_page_count(paths[0])
                # Обновляем max_value для start и end
                dpg.configure_item("start", max_value=total_pages)
                dpg.configure_item("end", max_value=total_pages)
                dpg.set_value("end", total_pages)

                data_numbers["end"]=total_pages

                # Если текущее значение end больше total_pages — обрезаем
                current_end = dpg.get_value("end")
                if current_end > total_pages:
                    dpg.set_value("end", total_pages)
                # То же для start (на всякий случай)
                current_start = dpg.get_value("start")
                if current_start > total_pages:
                    dpg.set_value("start", total_pages)

        except Exception as ex:
            error_str = str(ex)
            dpg.set_value("error_text", error_str)

    def select_file(name, id, only_directory=False):
        try:
            with dpg.file_dialog(label="Проводник", width=600, height=400, show=False,
                                 directory_selector=only_directory, callback=path_extractor, user_data=id,
                                 tag=f"filedialog_{id}"):
                dpg.add_file_extension(".*", color=(255, 255, 255, 255))
                dpg.add_file_extension(".xlsx", color=(40, 199, 39, 255), custom_text="EXCEL")
                dpg.add_file_extension(".pdf", color=(199, 39, 39, 255), custom_text="PDF")
                dpg.add_file_extension(".py", color=(0, 255, 255, 255))
                # dpg.add_button(label="Button on file dialog")
            with dpg.group(horizontal=True):
                dpg.add_button(label="Выбрать", callback=lambda: dpg.show_item(f"filedialog_{id}"))
                dpg.add_input_text(hint=name, callback=callback, tag=f"input_text_{id}", width=690)
            return paths[id]
        except Exception as ex:
            error_str = str(ex)
            dpg.set_value("error_text", error_str)

    def callback(sender, app_data):
        print('OK was clicked.')
        print("Sender: ", sender)
        print("App Data: ", app_data)

    dpg.create_context()
    dpg.create_viewport(title='Modbus App', width=window_width, height=window_height, resizable=False)

    with dpg.font_registry():
        with dpg.font(f'C:\\Windows\\Fonts\\arial.ttf', 18, default_font=True, id="Default font"):
            dpg.add_font_range_hint(dpg.mvFontRangeHint_Cyrillic)
    dpg.bind_font("Default font")

    with dpg.window(label="Главное окно", pos=(0, 0), width=window_width, height=window_height, no_resize=True,
                    no_move=True, no_collapse=True, no_close=True, no_title_bar=True, no_bring_to_front_on_focus=True):

        with dpg.collapsing_header(label="1. Источник данных(КД)", leaf=True):
            with dpg.group(horizontal=True):
                dpg.add_text("Выберите файл с конструкторской документацией(КД):")
                dpg.add_text("(PDF)", color=(255, 0, 0))
            path_data = select_file("Путь к файлу с данными", 0)
            dpg.add_text("Выберите диапозон страниц ТОЛЬКО с таблицей сигналов в КД:")
            with dpg.group(horizontal=True):
                dpg.add_text("Начало с ")
                dpg.add_input_int(tag="start", width=100, callback=set_numbers, min_clamped=True, min_value=1,
                                  max_clamped=True, max_value=300, default_value=1)
                dpg.add_text(" до ")
                dpg.add_input_int(tag="end", width=100, callback=set_numbers, min_clamped=True, min_value=1,
                                  max_clamped=True, max_value=300, default_value=1)
                dpg.add_text("страницы(включительно).")
            dpg.add_text("Выбор имен обозначений типов каналов в КД:")
            with dpg.group(horizontal=True):
                dpg.add_combo(label="DO", width=150)
                dpg.add_combo(label="DI", width=150)
                dpg.add_combo(label="AO", width=150)
                dpg.add_combo(label="AI", width=150)
        with dpg.collapsing_header(label="2. Шаблон карты Modbus", leaf=True):
            with dpg.group(horizontal=True):
                dpg.add_text("Выберите файл XLSX с шаблонами карт:")
                dpg.add_text("(XLSX)", color=(0, 255, 0))
            path_sample = select_file("Путь к файлу с шаблоном", 1)
            with dpg.group(horizontal=True):
                dpg.add_text("Выберите шаблон из списка:")
                dpg.add_combo(width=200, callback=set_sample_name, tag="sample_combo")
                dpg.add_button(label="Считать названия шаблонов", callback=sample_list_extr, show=True,
                               tag="button_sample_list")

        with dpg.collapsing_header(label="3. Обработка и путь сохранения", leaf=True):
            dpg.add_text("В данную папку будет сохранен файл со списком сигналов(signals.xlsx)"
                         + " и файл карты регистров(modbus_map.xlsx).", wrap=550)
            path_export = select_file("Путь в папку с проектом", 2, True)

        with dpg.collapsing_header(label="4. Данные для визуализации", leaf=True, ):
            dpg.add_text("Выберите необходимые таблицы для визуализации Weintek EasyBuilderPro")
            with dpg.group(horizontal=True):
                dpg.add_checkbox(label="Адресные метки", )
                dpg.add_checkbox(label="Таблица строк")
                dpg.add_checkbox(label="Журнал")
                dpg.add_button(label="Сформировать таблицы")

        with dpg.group(horizontal=True):
            dpg.add_button(label="Сформировать список сигналов", callback=create_xlsx, pos=(8, window_height - 160),
                           width=250)
            dpg.add_loading_indicator(style=1, radius=1.3, color=(0, 0, 255), show=False, tag="load_indic_1")

        with dpg.group(horizontal=True):
            dpg.add_button(label="Сформировать карту регистров", callback=create_map, pos=(8, window_height - 130),
                           width=250)
            dpg.add_loading_indicator(style=1, radius=1.3, color=(0, 0, 255), show=False, tag="load_indic_2")
            dpg.add_progress_bar(tag="progress_bar_map", show=False, width=200, )


        dpg.add_input_text(readonly=True, pos=(8, window_height - 70), width=800 - 32, tag="error_text")

    with dpg.theme() as global_theme:

        with dpg.theme_component(dpg.mvAll):
            dpg.add_theme_color(dpg.mvThemeCol_FrameBg, (130, 112, 90), category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 5, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_TextDisabled, (255, 255, 255), category=dpg.mvThemeCat_Core)

        with dpg.theme_component(dpg.mvInputInt):
            dpg.add_theme_color(dpg.mvThemeCol_FrameBg, (114, 130, 90), category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 5, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_Text, (255, 255, 255), category=dpg.mvThemeCat_Core)

        with dpg.theme_component(dpg.mvButton):
            dpg.add_theme_color(dpg.mvThemeCol_Button, (143, 143, 143), category=dpg.mvThemeCat_Core)

    with dpg.theme() as item_theme:

        with dpg.theme_component(dpg.mvInputText):
            dpg.add_theme_color(dpg.mvThemeCol_FrameBg, (69, 69, 69), category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 5, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_Text, (196, 135, 135), category=dpg.mvThemeCat_Core)

    dpg.bind_theme(global_theme)
    dpg.bind_item_theme("error_text", item_theme)

    dpg.setup_dearpygui()
    dpg.show_viewport()
    dpg.start_dearpygui()
    dpg.destroy_context()


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    gui()
    # pdf_to_xlsx("pdf_test2.pdf","xlsx_test.xlsx", 38, 45)
    # create_modbus_map("Карта регистров (шаблон).xlsx", "test_modbus.xlsx", "Шаблон 2", "xlsx_test.xlsx")
