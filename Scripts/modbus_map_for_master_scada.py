import openpyxl

def convert_modbus_map(path_map="",name_new_map="modbus_for_master_sascada", name_sheet = "Шаблон 2"):
    try:
        workbook = openpyxl.load_workbook(path_map)
    except Exception as e:
        print(f"Ошибка при открытии файла: {e}")
        return False

    if name_sheet not in workbook.sheetnames:
        print(f"Лист '{name_sheet}' не найден в таблице")
        return False

    ws_data = workbook[name_sheet]

    # Именованные столбцы для ясности
    ADDR_LABEL = 'K'  # Адресная метка
    DEVICE_NAME = 'J'  # Имя устройства
    BITMAP_TYPE = 'L'  # Тип: BITMAP или нет
    DATA_TYPE = 'D'  # Тип данных (BOOL и др.)
    REG_ADDR = 'B'  # Адрес регистра
    SIGNAL_NAME = 'A'  # Имя сигнала

    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Sheet 1"

    name_plc = "PLC"
    row_counter = 1

    def convert_data_type(data_type: str) -> str:
        print(data_type)
        mapping = {
            "FLOAT(4 byte)": "REAL",
            "WORD(2 byte)": "WORD",
            # Можно добавить другие типы при необходимости
        }
        return mapping.get(data_type.strip(), "Неопределенный")

    max_row = ws_data.max_row
    for row in range(2, max_row + 1):
        addr_label_val = ws_data[f"{ADDR_LABEL}{row}"].value
        if not addr_label_val:  # Пропускаем пустые строки
            continue

        device_name_val = ws_data[f"{DEVICE_NAME}{row}"].value or ""
        reg_addr_val = ws_data[f"{REG_ADDR}{row}"].value
        data_type_val = ws_data[f"{DATA_TYPE}{row}"].value
        panel_data_type = convert_data_type(data_type_val)
        signal_name_val = ws_data[f"{SIGNAL_NAME}{row}"].value

        # Определяем тип функции
        func_type = "HOLDING_REGISTERS"

        signal_full_name = f"{addr_label_val}_{device_name_val}" if device_name_val else addr_label_val

        new_sheet.cell(row=row_counter, column=1, value=signal_full_name)
        new_sheet.cell(row=row_counter, column=2, value=name_plc)
        new_sheet.cell(row=row_counter, column=3, value=func_type)
        new_sheet.cell(row=row_counter, column=4, value=reg_addr_val)
        new_sheet.cell(row=row_counter, column=5, value=signal_name_val)
        new_sheet.cell(row=row_counter, column=6, value=panel_data_type)

        row_counter += 1

    # Сохранение
    try:
        new_workbook.save(name_new_map + ".xlsx")  # или просто name_new_map, если уже с .xlsx
        print(f"Файл успешно сохранён как {name_new_map}.xlsx")
        return True
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")
        return False



if __name__ == "__main__" :
    convert_modbus_map("modbus_map.xlsx", name_sheet="Шаблон")