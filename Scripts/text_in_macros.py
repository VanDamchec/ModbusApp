import openpyxl
import os
import argparse

ALLOWED_SIGNALS = {"AI", "AO", "DI", "DO", "MTR", "MTRPID", "DVLV", "AVLV"}
MAX_LENGTH = 20  # Максимальная длина строки для String2Unicode

def safe_truncate(text, max_len=MAX_LENGTH):
    """Обрезает текст до max_len символов. Не добавляет '...', чтобы не тратить символы."""
    return text[:max_len] if len(text) > max_len else text

def generate_all_case_files(input_excel, output_dir="output", output_file_base="output_code", max_len=MAX_LENGTH):
    if not os.path.exists(input_excel):
        print(f"Файл {input_excel} не найден.")
        return

    try:
        wb = openpyxl.load_workbook(input_excel, read_only=True)
    except Exception as e:
        print(f"Ошибка при открытии Excel-файла: {e}")
        return

    os.makedirs(output_dir, exist_ok=True)
    processed_count = 0

    for sheet_name in wb.sheetnames:
        if sheet_name not in ALLOWED_SIGNALS:
            print(f"Пропускаем лист '{sheet_name}' — не в списке допустимых сигналов.")
            continue

        ws = wb[sheet_name]
        entries = []
        for row in ws.iter_rows(min_col=1, max_col=2, min_row=2):
            name = row[0].value
            desc = row[1].value if len(row) > 1 else None
            if name is not None:
                entries.append((str(name), str(desc) if desc is not None else ""))

        if not entries:
            print(f"Лист '{sheet_name}' пуст (столбец A). Файл не будет создан.")
            continue

        file_name = os.path.join(output_dir, f"{output_file_base}_{sheet_name}.txt")
        try:
            with open(file_name, "w", encoding="utf-8") as f:
                for i, (name, desc) in enumerate(entries, 1):
                    # Подготовка строки для String2Unicode
                    name_clean = name.strip()
                    desc_clean = desc.strip()

                    if desc_clean:
                        candidate = f"{name_clean}. {desc_clean}"
                    else:
                        candidate = name_clean

                    if len(candidate) <= max_len:
                        use_text = candidate
                        comment_desc = None
                    else:
                        # Превышает лимит — используем только имя (обрезанное при необходимости)
                        use_text = safe_truncate(name_clean, max_len)
                        comment_desc = desc_clean if desc_clean else None

                        # Запись комментария (если нужно и есть что писать)
                        f.write(f"// {comment_desc} (Тег+Описание длинее {max_len} символов )\n" )

                    # Запись case-блока
                    f.write(f"case {i}\n")
                    f.write(f'    String2Unicode("{use_text}", unicode_str[0])\n')
                    f.write(f"    break\n\n")

            print(f"Создан файл: {file_name} с {len(entries)} case-ами.")
            processed_count += 1
        except Exception as e:
            print(f"Ошибка при записи файла {file_name}: {e}")

    if processed_count == 0:
        print("Ни один допустимый лист с данными не найден.")
    else:
        print(f"Всего создано файлов: {processed_count}")
    return True


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Генерация case-блоков с поддержкой описаний (макс. 20 символов в строке)."
    )
    parser.add_argument(
        "input_excel",
        help="Путь к входному Excel-файлу (например, signals.xlsx)"
    )
    parser.add_argument(
        "--output_dir",
        default="output",
        help="Имя папки для сохранения выходных файлов (по умолчанию: output)"
    )
    parser.add_argument(
        "--output_file",
        default="output_code",
        help="Базовое имя выходного файла (по умолчанию: output_code)"
    )

    parser.add_argument(
        "--len",
        type=int,
        default=40,
        help="Максимальная длина строки в символах"
    )

    args = parser.parse_args()

    generate_all_case_files(
        input_excel=args.input_excel,
        output_dir=args.output_dir,
        output_file_base=args.output_file,
        max_len=args.len
    )