import openpyxl
import xlwt
import os

def xlsx_to_xls(xlsx_path, xls_path):
    """
    Преобразует файл Excel .xlsx в .xls.

    Args:
        xlsx_path (str): Путь к входному файлу .xlsx.
        xls_path (str): Путь, куда сохранить выходной файл .xls.

    Returns:
        bool: True, если преобразование прошло успешно, иначе False.
    """
    try:
        # Проверяем, существует ли входной файл
        if not os.path.exists(xlsx_path):
            print(f"Ошибка: Входной файл '{xlsx_path}' не найден.")
            return False

        # Загружаем книгу .xlsx
        wb_xlsx = openpyxl.load_workbook(xlsx_path, data_only=True) # data_only=True, чтобы получить вычисленные значения, а не формулы
        sheet_names = wb_xlsx.sheetnames

        # Создаем новую книгу .xls
        wb_xls = xlwt.Workbook(encoding='utf-8')

        for sheet_name in sheet_names:
            ws_xlsx = wb_xlsx[sheet_name]
            ws_xls = wb_xls.add_sheet(sheet_name)

            # Ограничиваем количество строк/столбцов, так как .xls ограничен
            max_row_xls = min(ws_xlsx.max_row, 65536)
            max_col_xls = min(ws_xlsx.max_column, 256)

            print(f"Обработка листа '{sheet_name}': строки 1-{max_row_xls}, столбцы 1-{max_col_xls}")

            for row in range(1, max_row_xls + 1):
                for col in range(1, max_col_xls + 1):
                    cell_xlsx = ws_xlsx.cell(row=row, column=col)
                    value = cell_xlsx.value

                    # xlwt использует 0-based индексы
                    # Пытаемся записать значение. xlwt не поддерживает все типы данных .xlsx напрямую,
                    # но для большинства простых значений (str, int, float, bool, None) это работает.
                    # Стили и формулы не переносятся.
                    try:
                        # xlwt не любит datetime, timedelta и некоторые другие типы.
                        # Преобразуем их в строки, если нужно.
                        if isinstance(value, (int, float)):
                            ws_xls.write(row - 1, col - 1, value)
                        elif isinstance(value, (str, bool)) or value is None:
                            # xlwt может записать bool как 1/0, str как есть, None как пустую ячейку
                            ws_xls.write(row - 1, col - 1, value)
                        else:
                            # Для других типов (например, datetime) преобразуем в строку
                            ws_xls.write(row - 1, col - 1, str(value))
                    except Exception as e:
                        # Если возникла ошибка при записи (например, тип данных xlrd не поддерживает),
                        # записываем как строку или пустую ячейку
                        print(f"Предупреждение: Не удалось записать значение {value} (тип {type(value)}) в ячейку {xlwt.formula.cell(row - 1, col - 1)} листа '{sheet_name}': {e}")
                        # Попробуем записать как строку
                        try:
                            ws_xls.write(row - 1, col - 1, str(value))
                        except:
                             # Если и строку не удалось, оставим пустой
                             ws_xls.write(row - 1, col - 1, "")

        # Сохраняем книгу .xls
        wb_xls.save(xls_path)
        print(f"Файл '{xls_path}' успешно создан из '{xlsx_path}'.")
        wb_xlsx.close() # Закрываем .xlsx файл после использования
        return True

    except Exception as e:
        print(f"Ошибка при преобразовании '{xlsx_path}' в '{xls_path}': {e}")
        return False

# --- Пример использования ---
if __name__ == "__main__":
    input_xlsx_file = "output_alarms.xlsx" # Входной файл .xlsx
    output_xls_file = "output_alarms.xls"  # Выходной файл .xls

    success = xlsx_to_xls(input_xlsx_file, output_xls_file)
    if success:
        print("Преобразование завершено успешно.")
    else:
        print("Преобразование не удалось.")
