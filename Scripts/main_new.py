import pdfplumber as pl
import re
from collections import namedtuple
import pandas as pd
import openpyxl
from copy import copy
import dearpygui.dearpygui as dpg
import os
import tkinter as tk
from tkinter import filedialog

from text_table_generate import create_name_table
from modbus_map_for_panel import convert_modbus_map


def pdf_to_xlsx(path_pdf, path_xlsx, start_page, end_page):
    di_map = namedtuple("di_map", "poz name module_name module_poz channel contact type_signal".split())
    ai_map = namedtuple("ai_map", "poz name module_name module_poz channel contact type_signal".split())
    do_map = namedtuple("do_map", "poz name module_name module_poz channel contact type_signal".split())
    ao_map = namedtuple("ao_map", "poz name module_name module_poz channel contact type_signal".split())
    maps = [di_map, ai_map, do_map, ao_map]
    names = ["DI", "AI", "DO", "AO"]
    tables = [list() for _ in range(len(maps))]

    im_map = namedtuple("im_map", "poz name".split())
    im_names = ["DVLV", "AVLV", "MTR", "MTRPID", "BL"]
    im_tables = [list() for _ in range(len(im_names))]
    im_tables[0].append(im_map("Поз1", "Задвижка 1"))
    im_tables[1].append(im_map("Поз1", "Регулируемый клапан 1"))
    im_tables[2].append(im_map("Поз1", "Агрегат 1"))
    im_tables[3].append(im_map("Поз1", "Регулируемый мотор 1"))
    im_tables[4].append(im_map("Поз1", "Блокировка 1"))

    with pl.open(path_pdf) as pdf:
        pages = pdf.pages
        for page in pages[(start_page - 1):end_page]:
            table = page.extract_table()
            if not table:
                continue
            for line in table:
                line = [x for x in line if x is not None]
                if not line:
                    continue
                for i in range(4):
                    if re.search(f"{names[i]}" + r"\d{1,2}", str(line)):
                        if re.search(r"№", line[0]) and len(line) > 12:
                            del line[:2]

                        poz = line[0]
                        name_param = line[1].replace("\n", " ")
                        module_name = line[-1].replace("\n", " ")
                        module_poz = line[-3]
                        channel = line[-2]

                        if len(line) == 18:
                            type_sig = line[-9]
                            contact = line[-7]
                        else:
                            type_sig = line[-8]
                            contact = line[-6]

                        if re.search(r"резерв", poz, flags=re.IGNORECASE) or re.search(r"резерв", name_param,
                                                                                       flags=re.IGNORECASE):
                            poz = "РЕЗЕРВ"
                            name_param = " "
                        tables[i].append(maps[i](poz, name_param, module_name, module_poz, channel, contact,
                                                 type_sig))

        with pd.ExcelWriter(path_xlsx) as writer:
            for i in range(len(maps)):
                df = pd.DataFrame(tables[i])
                df.to_excel(writer, sheet_name=names[i], index=False)
            for i in range(len(im_tables)):
                df = pd.DataFrame(im_tables[i])
                df.to_excel(writer, sheet_name=im_names[i], index=False)
    return True


def data_find(data_path):
    IO_num = {}
    wb_data = openpyxl.load_workbook(data_path)
    for sheet in wb_data.sheetnames:
        ws_data = wb_data[sheet]
        count = ws_data.max_row - 2
        IO_num.update([(sheet, count)])
    return IO_num


def marker_find(sample_path, name_sheet):
    markers = []
    size = []
    start_adress = []
    number_cell_marker = -1
    number_row = 0

    wb_sample = openpyxl.load_workbook(sample_path, data_only=True)
    ws_sample = wb_sample[name_sheet]

    for row in ws_sample.rows:
        if number_cell_marker == -1:
            for cell in row:
                if cell.value == "МЕТКА":
                    number_cell_marker = cell.column
                    break
        else:
            marker_cell = row[number_cell_marker - 1]
            marker = marker_cell.value
            if marker is not None:
                markers.append(marker)
                current_row_num = row[1].row
                size.append(current_row_num - number_row - 1)
                number_row = current_row_num
                if number_cell_marker < len(row):
                    start_adress.append(row[number_cell_marker].value)
                else:
                    start_adress.append(None)

    if markers:
        markers = markers[:-1]
    if size:
        size = size[1:]

    return markers, size, start_adress


def create_modbus_map(sample_map_path, out_map_path, name_sheet,
                      data_path, show_progress=False, progress_bar_tag="",
                      delete_other_sheets=False):
    size_val_modbus = 1
    bit_adress_modbus = 0
    bit_modbus = 0

    IO_num = data_find(data_path)
    markers, size, start_modbus_adress = marker_find(sample_map_path, name_sheet)
    num_channel = [IO_num.get(marker.split("_")[0]) for marker in markers if IO_num.get(marker.split("_")[0]) is not None]

    wb_data = openpyxl.load_workbook(data_path)
    wb_sample = openpyxl.load_workbook(sample_map_path)
    list_all_ws = wb_sample.sheetnames

    if delete_other_sheets:
        for item in list_all_ws:
            if item != name_sheet:
                wb_sample.remove(wb_sample[item])

    ws_sample = wb_sample[name_sheet]
    k = 0
    for marker in markers:
        type_channel = marker.split("_")[0]
        list_all_ws = wb_data.sheetnames
        if type_channel not in list_all_ws:
            continue
        ws_data = wb_data[type_channel]
        row_num_data = {"AI": 2, "AO": 2, "DI": 2, "DO": 2, "DVLV": 2, "AVLV": 2, "MTR": 2,
                        "BL": 2, "MTRPID": 2}

        bit_adress_modbus = 0
        row_num = 0
        for row in ws_sample.rows:
            row_num += 1
            if row[8].value == marker:
                # === Сохраняем объединения до вставки ===
                merged_ranges_before = list(ws_sample.merged_cells.ranges)

                # === Вставляем строки ===
                ws_sample.insert_rows(row_num + 1 + size[k], num_channel[k] * size[k])

                # === Восстанавливаем объединения, исключая новые строки ===
                ws_sample.merged_cells.ranges.clear()

                for merged_range in merged_ranges_before:
                    min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col

                    # Если объединение начинается **до вставки**, и заканчивается **до вставки** — оставляем как есть
                    if max_row < row_num + 1 + size[k]:
                        ws_sample.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row,
                                              end_column=max_col)
                    # Если объединение **начинается до** и **пересекает** вставку — расширяем
                    elif min_row < row_num + 1 + size[k] <= max_row:
                        new_max_row = max_row + num_channel[k] * size[k]
                        ws_sample.merge_cells(start_row=min_row, start_column=min_col, end_row=new_max_row,
                                              end_column=max_col)
                    # Если объединение **ниже вставки** — смещаем вниз
                    elif min_row >= row_num + 1 + size[k]:
                        new_min_row = min_row + num_channel[k] * size[k]
                        new_max_row = max_row + num_channel[k] * size[k]
                        ws_sample.merge_cells(start_row=new_min_row, start_column=min_col, end_row=new_max_row,
                                              end_column=max_col)

                for n in range(num_channel[k] + 1):
                    for j in range(size[k] * n + row_num + 1,
                                   size[k] * (n + 1) + row_num + 1):

                        ws_sample.cell(row=j, column=10).value = f"{type_channel}{n + 1}"

                        if n != num_channel[k]:
                            for col in [11, 12]:
                                old_cell = ws_sample.cell(row=j, column=col)
                                new_cell = ws_sample.cell(row=j + size[k], column=col)

                                is_merged = any(new_cell.coordinate in merged_range for merged_range in
                                                ws_sample.merged_cells.ranges)
                                if not is_merged:
                                    new_cell.value = old_cell.value
                                    new_cell._style = copy(old_cell._style)
                                    if old_cell.hyperlink:
                                        new_cell.hyperlink = old_cell.hyperlink

                            for i in range(1, 9):
                                old_cell = ws_sample.cell(row=j, column=i)
                                new_cell = ws_sample.cell(row=j + size[k], column=i)

                                # Проверяем, не является ли новая ячейка частью объединённого диапазона
                                is_merged = any(new_cell.coordinate in merged_range for merged_range in
                                                ws_sample.merged_cells.ranges)
                                if not is_merged:
                                    new_cell.value = old_cell.value
                                    new_cell._style = copy(old_cell._style)
                                    if old_cell.hyperlink:
                                        new_cell.hyperlink = old_cell.hyperlink

                        for i in range(1, 9):
                            new_cell = ws_sample.cell(row=j, column=i)
                            is_merged = any(
                                new_cell.coordinate in merged_range for merged_range in ws_sample.merged_cells.ranges)
                            if not is_merged:
                                if i == 1:
                                    old_value = str(new_cell.value)
                                    new_value = (str(ws_data.cell(row=row_num_data[type_channel], column=1).value) +
                                                 "-" + str(
                                                ws_data.cell(row=row_num_data[type_channel], column=2).value))
                                    new_value = old_value.replace("$$", new_value)
                                    new_cell.value = new_value
                                elif i == 2:
                                    ws_sample.cell(row=j,
                                                   column=2).value = f"{start_modbus_adress[k]}.{bit_adress_modbus}"
                                    old_cell_data_type = str(ws_sample.cell(row=j, column=4).value)
                                    new_cell_data_type = str(
                                        ws_sample.cell(row=j + 1, column=4).value) if j + 1 <= ws_sample.max_row else ""
                                    if "4 byte" in old_cell_data_type:
                                        size_val_modbus = 2
                                        bit_modbus = 0
                                        if "BOOL" in new_cell_data_type:
                                            start_modbus_adress[k] += size_val_modbus
                                        ws_sample.cell(row=j, column=2).value = f"{start_modbus_adress[k]}"
                                    elif "2 byte" in old_cell_data_type:
                                        size_val_modbus = 1
                                        bit_modbus = 0
                                        if "BOOL" in new_cell_data_type:
                                            start_modbus_adress[k] += size_val_modbus
                                        ws_sample.cell(row=j, column=2).value = f"{start_modbus_adress[k]}"
                                    elif "BOOL" in old_cell_data_type:
                                        size_val_modbus = 0
                                        bit_modbus = 1
                                        if "4 byte" in new_cell_data_type:
                                            size_val_modbus = -1
                                            ws_sample.cell(row=j,
                                                           column=2).value = f"{start_modbus_adress[k]}.{bit_adress_modbus}"
                                    else:
                                        ws_sample.cell(row=j,
                                                       column=2).value = f"{start_modbus_adress[k]}.{bit_adress_modbus}"
                                elif i == 6:
                                    channel_name = (f"{ws_data.cell(row=row_num_data[type_channel], column=5).value}" +
                                                    f"({ws_data.cell(row=row_num_data[type_channel], column=4).value})")
                                    ws_sample.cell(row=j, column=i).value = channel_name

                        start_modbus_adress[k] += size_val_modbus
                        bit_adress_modbus += bit_modbus

                        if ((size[k] > 1 and bit_adress_modbus >= (4 * size[k])) or
                                bit_adress_modbus > 15 or
                                (type_channel == "DVLV" and bit_adress_modbus >= size[k]) or
                                (type_channel == "AI" and bit_adress_modbus >= size[k]) or
                                (16 // size[k] == 1 and bit_adress_modbus >= size[k])):
                            start_modbus_adress[k] += 1
                            bit_adress_modbus = 0

                    row_num_data[type_channel] += 1

        k += 1
        progress = round(k / len(markers) * 100, 1)
        if show_progress:
            dpg.set_value(progress_bar_tag, k / len(markers))
            dpg.configure_item(progress_bar_tag, show=True, overlay=f"{progress}%")

    wb_sample.save(out_map_path)
    return True


# ============ GUI ============

global path_data
global path_sample
global path_export
global paths
paths = ["", "", ""]
global sample_list
global sample_name
global data_numbers
data_numbers = {"start": 1, "end": 1}
global error_str
error_str = ""
global sample_sheetnames
sample_sheetnames = []
global generate_address_labels, generate_name_table, generate_journal
generate_address_labels = False
generate_name_table = False
generate_journal = False  # пока не используется, но для будущего


def handle_overwrite_response(filename, on_confirm, on_cancel):
    """Показывает диалог перезаписи и вызывает колбэк после выбора."""
    def on_yes():
        dpg.delete_item("overwrite_modal")
        on_confirm(filename)

    def on_no():
        dpg.delete_item("overwrite_modal")
        on_cancel()

    def on_rename(sender, app_data):
        nonlocal filename
        new_name = app_data.strip()
        if new_name:
            filename = os.path.join(os.path.dirname(filename), new_name)

    with dpg.mutex():
        with dpg.window(label="Файл существует", modal=True, no_close=True, tag="overwrite_modal"):
            dpg.add_text(f"Файл {os.path.basename(filename)} уже существует.")
            dpg.add_text("Перезаписать?")
            dpg.add_input_text(default_value=os.path.basename(filename), callback=on_rename, width=300, tag="rename_input")
            with dpg.group(horizontal=True):
                dpg.add_button(label="Да", callback=on_yes)
                dpg.add_button(label="Нет", callback=on_no)

def create_map(sender, appdata):
    global sample_name
    dpg.configure_item("load_indic_2", show=True)
    try:
        if not paths[1]:
            raise ValueError("Не выбран файл шаблона карты Modbus")
        if not sample_name:
            raise ValueError("Не выбран лист шаблона")
        if sample_name not in sample_sheetnames:
            raise ValueError(f"Выбранный лист '{sample_name}' отсутствует в файле шаблона. Перезагрузите шаблон.")

        if not paths[2]:
            raise ValueError("Не указан путь для сохранения")

        signals_path = os.path.join(paths[2], "signals.xlsx")
        if not os.path.exists(signals_path):
            raise FileNotFoundError("Файл signals.xlsx не найден. Сначала сформируйте список сигналов.")

        out_map_path = os.path.join(paths[2], "modbus_map.xlsx")

        def continue_with_path(final_path):
            try:
                create_modbus_map(paths[1], final_path, sample_name, signals_path,
                                  show_progress=True, progress_bar_tag="progress_bar_map")
                dpg.set_value("error_text", f"Карта сохранена: {final_path}")
            except Exception as ex:
                dpg.set_value("error_text", str(ex))
            finally:
                dpg.configure_item("load_indic_2", show=False)

        def on_cancel():
            dpg.set_value("error_text", "Операция отменена пользователем.")
            dpg.configure_item("load_indic_2", show=False)

        if os.path.exists(out_map_path):
            handle_overwrite_response(out_map_path, continue_with_path, on_cancel)
        else:
            continue_with_path(out_map_path)

    except Exception as ex:
        dpg.set_value("error_text", str(ex))
        dpg.configure_item("load_indic_2", show=False)

def create_xlsx(sender, app_data):
    dpg.configure_item("load_indic_1", show=True)
    try:
        if not paths[0]:
            raise ValueError("Не выбран файл КД")
        if not paths[0].lower().endswith('.pdf'):
            raise ValueError("Файл КД должен быть в формате PDF")
        if not paths[2]:
            raise ValueError("Не указан путь для сохранения")

        signals_path = os.path.join(paths[2], "signals.xlsx")

        # Функция продолжения после выбора
        def continue_with_path(final_path):
            try:
                pdf_to_xlsx(paths[0], final_path, data_numbers["start"], data_numbers["end"])
                dpg.set_value("error_text", f"Сигналы сохранены: {final_path}")
            except Exception as ex:
                dpg.set_value("error_text", str(ex))
            finally:
                dpg.configure_item("load_indic_1", show=False)

        def on_cancel():
            dpg.set_value("error_text", "Операция отменена пользователем.")
            dpg.configure_item("load_indic_1", show=False)

        if os.path.exists(signals_path):
            handle_overwrite_response(signals_path, continue_with_path, on_cancel)
        else:
            continue_with_path(signals_path)

    except Exception as ex:
        dpg.set_value("error_text", str(ex))
        dpg.configure_item("load_indic_1", show=False)

def create_table(sender, appdata):
    try:
        if not paths[2]:
            raise ValueError("Не указан путь для сохранения")

        # Сначала показываем диалог выбора файлов
        modbus_map_path = select_file_system_dialog(
            filetypes=[("Excel файлы", "*.xlsx")],
            initial_dir=paths[2],
            title="Выберите файл modbus_map.xlsx"
        )
        if not modbus_map_path:
            return  # пользователь отменил

        signals_path = select_file_system_dialog(
            filetypes=[("Excel файлы", "*.xlsx")],
            initial_dir=paths[2],
            title="Выберите файл signals.xlsx"
        )
        if not signals_path:
            return  # пользователь отменил

        if not os.path.exists(signals_path):
            raise FileNotFoundError("Файл signals.xlsx не найден.")
        if not os.path.exists(modbus_map_path):
            raise FileNotFoundError("Файл modbus_map.xlsx не найден.")

        # Загружаем список листов из modbus_map.xlsx
        wb = openpyxl.load_workbook(modbus_map_path, read_only=True)
        sheetnames = wb.sheetnames
        wb.close()

        def start_generation(selected_sheet_name):
            # Теперь можно запускать генерацию
            sample_name = selected_sheet_name  # используем выбранный лист

            # --- Логика генерации ---
            messages = []

            if not (generate_address_labels or generate_name_table or generate_journal):
                raise ValueError("Не выбраны таблицы для генерации")

            # 1. Таблица строк
            if generate_name_table:
                try:
                    create_name_table(
                        out_map_path=paths[2],
                        data_path=signals_path,
                        file_name="name_table.xlsx",
                        name_order=list(range(0, 50)),
                        start_id=0,
                        enable_export=True
                    )
                    messages.append("Таблица строк создана")
                except PermissionError:
                    raise PermissionError("Файл name_table.xlsx уже открыт в Excel. Закройте его и повторите попытку.")

            # 2. Адресные метки
            if generate_address_labels:
                try:
                    success = convert_modbus_map(
                        path_map=modbus_map_path,
                        name_new_map=os.path.join(paths[2], "modbus_for_panel"),
                        name_sheet=sample_name
                    )
                    if success:
                        messages.append("Адресные метки созданы")
                    else:
                        raise RuntimeError("Ошибка при создании адресных меток")
                except PermissionError:
                    raise PermissionError("Файл modbus_for_panel.xlsx уже открыт в Excel. Закройте его и повторите попытку.")

            if generate_journal:
                messages.append("Журнал: не реализован")

            dpg.set_value("error_text", "; ".join(messages) if messages else "Ничего не сгенерировано")

        # Показываем диалог выбора листа
        show_sheet_selection_dialog_for_generation(sheetnames, start_generation)

    except Exception as ex:
        dpg.set_value("error_text", f"Ошибка: {str(ex)}")

def set_numbers(sender, app_data):
    data_numbers[sender] = app_data

def set_sample_name(sender, app_data):
    global sample_name
    sample_name = str(app_data)

def show_sheet_selection_dialog(sheetnames):
    """Показывает модальное окно для выбора листа из списка."""
    def on_select(sender, app_data):
        global sample_name
        selected = dpg.get_value("sheet_selector_combo")
        if selected:
            sample_name = selected
            dpg.configure_item("sample_combo", items=sheetnames)
            dpg.set_value("sample_combo", selected)
            dpg.set_value("error_text", f"Выбран лист: {selected}")
        dpg.delete_item("sheet_selection_modal")

    def on_cancel():
        dpg.delete_item("sheet_selection_modal")
        dpg.set_value("error_text", "Не выбран лист шаблона. Выберите файл снова.")

    with dpg.mutex():
        with dpg.window(label="Выберите лист шаблона", modal=True, no_close=True, tag="sheet_selection_modal"):
            dpg.add_text("Лист 'Шаблон' не найден. Выберите нужный лист:")
            dpg.add_combo(items=sheetnames, default_value=sheetnames[0] if sheetnames else "", tag="sheet_selector_combo", width=250)
            with dpg.group(horizontal=True):
                dpg.add_button(label="Выбрать", callback=on_select)
                dpg.add_button(label="Отмена", callback=on_cancel)

def sample_list_extr(sender, app_data):
    global sample_list, sample_name, sample_sheetnames
    try:
        path_sample = paths[1]
        if not path_sample.lower().endswith('.xlsx'):
            raise ValueError("Файл шаблона должен быть в формате XLSX")
        wb_sample = openpyxl.load_workbook(path_sample, read_only=True)
        sheetnames = wb_sample.sheetnames
        wb_sample.close()
        sample_sheetnames = sheetnames  # Сохраняем актуальный список

        if "Шаблон" in sheetnames:
            sample_name = "Шаблон"
            dpg.configure_item("sample_combo", items=sheetnames)
            dpg.set_value("sample_combo", "Шаблон")
            dpg.set_value("error_text", "Автоматически выбран лист: Шаблон")
        else:
            # Показываем диалог выбора листа
            show_sheet_selection_dialog(sheetnames)
    except Exception as ex:
        dpg.set_value("error_text", str(ex))

def get_pdf_page_count(pdf_path):
    try:
        with pl.open(pdf_path) as pdf:
            return len(pdf.pages)
    except Exception as ex:
        print(f"Ошибка при чтении PDF: {ex}")
        return 1

def path_extractor(sender, app_data, id):
    global sample_name, sample_sheetnames
    try:
        if sender == "filedialog_2":
            paths[id] = app_data["current_path"]
        else:
            paths[id] = str(list(app_data["selections"].values())[0])
        dpg.set_value(f"input_text_{id}", paths[id])

        # === Обработка PDF (id=0) ===
        if id == 0:
            if not paths[0].lower().endswith('.pdf'):
                dpg.set_value("error_text", "Файл КД должен иметь расширение .pdf")
                paths[0] = ""  # сбрасываем путь
                dpg.set_value("input_text_0", "")
                return
            # Только если это PDF — получаем количество страниц
            total_pages = get_pdf_page_count(paths[0])
            dpg.configure_item("start", max_value=total_pages)
            dpg.configure_item("end", max_value=total_pages)
            dpg.set_value("end", total_pages)
            data_numbers["end"] = total_pages

        # === Обработка XLSX шаблона (id=1) ===
        elif id == 1:
            sample_name = ""
            sample_sheetnames = []
            if not paths[1].lower().endswith('.xlsx'):
                dpg.set_value("error_text", "Файл шаблона должен иметь расширение .xlsx")
                paths[1] = ""  # сбрасываем путь
                dpg.set_value("input_text_1", "")
                return
            # Автоматически загружаем листы
            sample_list_extr(None, None)

        if id == 2:  # путь сохранения
            dpg.configure_item("btn_generate_tables", enabled=bool(paths[2]))

    except Exception as ex:
        dpg.set_value("error_text", str(ex))

def select_file(name, id, only_directory=False):
    global sample_name, sample_sheetnames
    try:
        # Определяем начальную директорию
        initial_dir = None
        if paths[id]:
            if only_directory:
                initial_dir = paths[id]
            else:
                # Если это файл, то извлекаем директорию
                if os.path.isfile(paths[id]):
                    initial_dir = os.path.dirname(paths[id])
                else:
                    initial_dir = paths[id]

        if only_directory:
            path = select_folder_system_dialog(initial_dir=initial_dir)
        else:
            if id == 0:  # файл КД (PDF)
                filetypes = [("PDF файлы", "*.pdf")]
            elif id == 1:  # файл шаблона (XLSX)
                filetypes = [("Excel файлы", "*.xlsx")]
            else:
                filetypes = [("Все файлы", "*.*")]
            path = select_file_system_dialog(
                filetypes=filetypes,
                initial_dir=initial_dir
            )

        if not path:
            return  # пользователь отменил

        # === Обработка PDF (id=0) ===
        if id == 0:
            if not path.lower().endswith('.pdf'):
                dpg.set_value("error_text", "Файл КД должен иметь расширение .pdf")
                return
            # Только если это PDF — получаем количество страниц
            total_pages = get_pdf_page_count(path)
            dpg.configure_item("start", max_value=total_pages)
            dpg.configure_item("end", max_value=total_pages)
            dpg.set_value("end", total_pages)
            data_numbers["end"] = total_pages

        # === Обработка XLSX шаблона (id=1) ===
        elif id == 1:
            sample_name = ""
            sample_sheetnames = []
            if not path.lower().endswith('.xlsx'):
                dpg.set_value("error_text", "Файл шаблона должен иметь расширение .xlsx")
                return
            # Автоматически загружаем листы
            paths[id] = path
            dpg.set_value(f"input_text_{id}", paths[id])
            sample_list_extr(None, None)
            return  # чтобы не сбрасывался путь ниже

        # === Сохраняем путь и обновляем интерфейс ===
        paths[id] = path
        dpg.set_value(f"input_text_{id}", paths[id])

        # === Обновляем кнопку генерации таблиц ===
        if id == 2:  # путь сохранения
            dpg.configure_item("btn_generate_tables", enabled=bool(paths[2]))

    except Exception as ex:
        dpg.set_value("error_text", str(ex))

def on_address_labels_checkbox(sender, app_data):
    if not paths[2]:
        dpg.set_value(sender, False)  # сбрасываем чекбокс
        dpg.set_value("error_text", "Сначала укажите путь для сохранения")
        return
    global generate_address_labels
    generate_address_labels = app_data

def on_name_table_checkbox(sender, app_data):
    if not paths[2]:
        dpg.set_value(sender, False)
        dpg.set_value("error_text", "Сначала укажите путь для сохранения")
        return
    global generate_name_table
    generate_name_table = app_data

def on_journal_checkbox(sender, app_data):
    if not paths[2]:
        dpg.set_value(sender, False)
        dpg.set_value("error_text", "Сначала укажите путь для сохранения")
        return
    global generate_journal
    generate_journal = app_data

def show_sheet_selection_dialog_for_generation(sheetnames, on_confirm):
    """Показывает диалог выбора листа и вызывает on_confirm с выбранным именем."""
    selected_sheet = [None]

    def on_select():
        selected = dpg.get_value("sheet_selector_combo_gen")
        if selected:
            selected_sheet[0] = selected
        dpg.delete_item("sheet_selection_modal_gen")
        if selected:
            on_confirm(selected)

    def on_cancel():
        dpg.delete_item("sheet_selection_modal_gen")

    with dpg.mutex():
        with dpg.window(label="Выберите лист из modbus_map.xlsx", modal=True, no_close=True, tag="sheet_selection_modal_gen"):
            dpg.add_text("Выберите лист, на основе которого генерировать таблицы:")
            dpg.add_combo(items=sheetnames, default_value=sheetnames[0] if sheetnames else "", tag="sheet_selector_combo_gen", width=250)
            with dpg.group(horizontal=True):
                dpg.add_button(label="ОК", callback=on_select)
                dpg.add_button(label="Отмена", callback=on_cancel)

def select_file_system_dialog(filetypes=None, initial_dir=None, title="Выберите файл"):
    """Открывает системный диалог для выбора файла."""
    root = tk.Tk()
    root.withdraw()  # скрываем главное окно
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes or [("Все файлы", "*.*")],
        initialdir=initial_dir
    )
    root.destroy()
    return file_path

def select_folder_system_dialog(initial_dir=None, title="Выберите папку"):
    """Открывает системный диалог для выбора папки."""
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(
        title=title,
        initialdir=initial_dir
    )
    root.destroy()
    return folder_path

def make_callback(name, id, only_directory):
    def callback(s, a):
        select_file(name, id, only_directory)
    return callback

def gui():
    window_width = 800
    window_height = 800

    dpg.create_context()
    dpg.create_viewport(title='Modbus App', width=window_width, height=window_height, resizable=False)

    with dpg.font_registry():
        with dpg.font('C:\\Windows\\Fonts\\arial.ttf', 18, default_font=True, id="Default font"):
            dpg.add_font_range_hint(dpg.mvFontRangeHint_Cyrillic)
    dpg.bind_font("Default font")

    with dpg.window(label="Главное окно", pos=(0, 0), width=window_width, height=window_height, no_resize=True,
                    no_move=True, no_collapse=True, no_close=True, no_title_bar=True, no_bring_to_front_on_focus=True):

        with dpg.collapsing_header(label="1. Источник данных(КД)", leaf=True):
            with dpg.group(horizontal=True):
                dpg.add_text("Выберите файл с конструкторской документацией(КД):")
                dpg.add_text("(PDF)", color=(255, 0, 0))
            with dpg.group(horizontal=True):
                dpg.add_button(label="Выбрать",
                               callback=make_callback( "Путь к файлу с данными", 0, False))
                dpg.add_input_text(hint="Путь к файлу с данными", tag="input_text_0", width=690)

            dpg.add_text("Выберите диапозон страниц ТОЛЬКО с таблицей сигналов в КД:")
            with dpg.group(horizontal=True):
                dpg.add_text("Начало с ")
                dpg.add_input_int(tag="start", width=100, callback=set_numbers, min_clamped=True, min_value=1,
                                  max_clamped=True, max_value=300, default_value=1)
                dpg.add_text(" до ")
                dpg.add_input_int(tag="end", width=100, callback=set_numbers, min_clamped=True, min_value=1,
                                  max_clamped=True, max_value=300, default_value=1)
                dpg.add_text("страницы(включительно).")

        with dpg.collapsing_header(label="2. Шаблон карты Modbus", leaf=True):
            with dpg.group(horizontal=True):
                dpg.add_text("Выберите файл XLSX с шаблонами карт:")
                dpg.add_text("(XLSX)", color=(0, 255, 0))
            with dpg.group(horizontal=True):
                dpg.add_button(label="Выбрать",
                               callback=make_callback("Путь к файлу с шаблоном", 1, False))
                dpg.add_input_text(hint="Путь к файлу с шаблоном", tag="input_text_1", width=690)
            with dpg.group(horizontal=True):
                dpg.add_text("Выберите шаблон из списка:")
                dpg.add_combo(width=200, callback=set_sample_name, tag="sample_combo")

        with dpg.collapsing_header(label="3. Обработка и путь сохранения", leaf=True):
            dpg.add_text("В данную папку будет сохранен файл со списком сигналов(signals.xlsx)"
                         + " и файл карты регистров(modbus_map.xlsx).", wrap=550)
            with dpg.group(horizontal=True):
                dpg.add_button(label="Выбрать",
                               callback=make_callback("Путь к папке с проектом", 2, True))
                dpg.add_input_text(hint="Путь к папке с проектом", tag="input_text_2", width=690)

        with dpg.collapsing_header(label="4. Данные для визуализации", leaf=True):
            dpg.add_text("Выберите необходимые таблицы для визуализации Weintek EasyBuilderPro")
            with dpg.group(horizontal=True):
                dpg.add_checkbox(label="Адресные метки", callback=on_address_labels_checkbox, tag="cb_address")
                dpg.add_checkbox(label="Таблица строк", callback=on_name_table_checkbox, tag="cb_name_table")
                dpg.add_checkbox(label="Журнал", callback=on_journal_checkbox, tag="cb_journal")
                dpg.add_button(label="Сформировать таблицы", callback=create_table, tag="btn_generate_tables")
                dpg.configure_item("btn_generate_tables", enabled=bool(paths[2]))

        with dpg.group(horizontal=True):
            dpg.add_button(label="Сформировать список сигналов", callback=create_xlsx, pos=(8, window_height - 160),
                           width=250)
            dpg.add_loading_indicator(style=1, radius=1.3, color=(0, 0, 255), show=False, tag="load_indic_1")

        with dpg.group(horizontal=True):
            dpg.add_button(label="Сформировать карту регистров", callback=create_map, pos=(8, window_height - 130),
                           width=250)
            dpg.add_loading_indicator(style=1, radius=1.3, color=(0, 0, 255), show=False, tag="load_indic_2")
            dpg.add_progress_bar(tag="progress_bar_map", show=False, width=200)

        dpg.add_input_text(readonly=True, pos=(8, window_height - 70), width=800 - 32, tag="error_text")

    with dpg.theme() as global_theme:
        with dpg.theme_component(dpg.mvAll):
            dpg.add_theme_color(dpg.mvThemeCol_FrameBg, (130, 112, 90), category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 5, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_TextDisabled, (255, 255, 255), category=dpg.mvThemeCat_Core)

        with dpg.theme_component(dpg.mvInputInt):
            dpg.add_theme_color(dpg.mvThemeCol_FrameBg, (114, 130, 90), category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 5, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_Text, (255, 255, 255), category=dpg.mvThemeCat_Core)

        with dpg.theme_component(dpg.mvButton):
            dpg.add_theme_color(dpg.mvThemeCol_Button, (143, 143, 143), category=dpg.mvThemeCat_Core)

    with dpg.theme() as item_theme:
        with dpg.theme_component(dpg.mvInputText):
            dpg.add_theme_color(dpg.mvThemeCol_FrameBg, (69, 69, 69), category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 5, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_Text, (196, 135, 135), category=dpg.mvThemeCat_Core)

    dpg.bind_theme(global_theme)
    dpg.bind_item_theme("error_text", item_theme)

    dpg.setup_dearpygui()
    dpg.show_viewport()
    dpg.start_dearpygui()
    dpg.destroy_context()


if __name__ == '__main__':
    gui()