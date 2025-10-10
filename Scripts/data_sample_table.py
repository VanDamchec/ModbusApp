import openpyxl
from openpyxl import Workbook
import sys
import os

def generate_sampling_table(input_file):
    # Проверяем, существует ли входной файл
    if not os.path.exists(input_file):
        print(f"Файл {input_file} не найден.")
        return

    # Открываем входной файл
    wb_input = openpyxl.load_workbook(input_file)
    if 'AI' not in wb_input.sheetnames:
        print("Лист 'AI' не найден в файле.")
        return

    ws_ai = wb_input['AI']
    signals = []

    # Считываем сигналы из первого столбца
    for row in ws_ai.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value:
            signals.append(str(cell.value).strip())

    # Создаем новую книгу и лист
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = "sheet1"

    # Записываем данные в лист
    ws_output.append(["Version: 4"])
    ws_output.append(["Data Sampling", "AI_PV"])
    ws_output.append([" ","Sample Mode","Time-based","1000 ms"])
    ws_output.append([" "," ", "High Priority: Off"])
    ws_output.append([" ","Read Address","ABAK PLC","4x","System Tag: Off","User-defined Tag: Off","1","IDX: null"])
    ws_output.append([" ","Data Record"])
    ws_output.append([" ", " ", signals[1], "32-bit Float","Left of decimal Pt. 4","Right of decimal Pt. 2",
                      "Leading zero Off","ABAK PLC","4x","System Tag: Off","User-defined Tag: Off","1","IDX: null",])


    #Добавляем сигналы
    for signal in signals[2:]:
        ws_output.append([
            " ",
            " ",
            signal,
            "32-bit Float",
            "Left of decimal Pt. 4",
            "Right of decimal Pt. 2",
            "Leading zero Off"
        ])
# Добавляем статичные параметры после сигналов
    ws_output.append([" ","History File","HMI memory"])
    ws_output.append([" "," ","Preservation Limit","31 day(s)/file(s)"])
    ws_output.append([" "," ","Sync Status Address: Off"])
    ws_output.append([" "," ","Folder Name","AI_PV"])
    ws_output.append([" "," ","Customized File","Automatic Mode"])
    ws_output.append([" "," "," ","","File Name","%Y%m%d"])
    ws_output.append([" "," "," ","Sort By","File name"])
    ws_output.append([" "," ","10000 Records limited: Off",])

    # Сохраняем результат
    output_file = os.path.splitext(input_file)[0] + "_out.xlsx"
    wb_output.save(output_file)

    print(f"Файл успешно сохранён: {output_file}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Пожалуйста, укажите путь к входному файлу Excel.")
    else:
        generate_sampling_table(sys.argv[1])