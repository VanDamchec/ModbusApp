from configparser import ConfigParser

import pdfplumber as pl
import re
from collections import namedtuple
import pandas as pd
import openpyxl
from copy import copy
import dearpygui.dearpygui as dpg
import os
import tkinter as tk
from tkinter import filedialog

from text_table_generate import create_name_table
from modbus_map_for_panel import convert_modbus_map
from data_sample_table import  generate_sampling_table
from text_in_macros import generate_all_case_files
from journal_map import generate_alarms_from_modbus_map

import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="pydevd_plugins")


def pdf_to_xlsx(path_pdf, path_xlsx, start_page, end_page):
    di_map = namedtuple("di_map", "poz name module_name module_poz channel contact type_signal".split())
    ai_map = namedtuple("ai_map", "poz name module_name module_poz channel contact type_signal".split())
    do_map = namedtuple("do_map", "poz name module_name module_poz channel contact type_signal".split())
    ao_map = namedtuple("ao_map", "poz name module_name module_poz channel contact type_signal".split())
    maps = [di_map, ai_map, do_map, ao_map]
    names = ["DI", "AI", "DO", "AO"]
    tables = [list() for _ in range(len(maps))]

    im_map = namedtuple("im_map", "poz name".split())
    im_names = ["DVLV", "AVLV", "MTR", "MTRPID", "BL"]
    im_tables = [list() for _ in range(len(im_names))]
    im_tables[0].append(im_map("Поз1", "Задвижка 1"))
    im_tables[1].append(im_map("Поз1", "Регулируемый клапан 1"))
    im_tables[2].append(im_map("Поз1", "Агрегат 1"))
    im_tables[3].append(im_map("Поз1", "Регулируемый мотор 1"))
    im_tables[4].append(im_map("Поз1", "Блокировка 1"))

    with pl.open(path_pdf) as pdf:
        pages = pdf.pages
        for page in pages[(start_page - 1):end_page]:
            table = page.extract_table()
            if not table:
                continue
            for line in table:
                line = [x for x in line if x is not None]
                if not line:
                    continue
                for i in range(4):
                    if re.search(f"{names[i]}" + r"\d{1,2}", str(line)):
                        if re.search(r"№", line[0]) and len(line) > 12:
                            del line[:2]

                        poz = line[0]
                        name_param = line[1].replace("\n", " ")
                        module_name = line[-1].replace("\n", " ")
                        module_poz = line[-3]
                        channel = line[-2]

                        if len(line) == 18:
                            type_sig = line[-9]
                            contact = line[-7]
                        else:
                            type_sig = line[-8]
                            contact = line[-6]

                        if re.search(r"резерв", poz, flags=re.IGNORECASE) or re.search(r"резерв", name_param,
                                                                                       flags=re.IGNORECASE):
                            poz = "РЕЗЕРВ"
                            name_param = " "
                        tables[i].append(maps[i](poz, name_param, module_name, module_poz, channel, contact,
                                                 type_sig))

        with pd.ExcelWriter(path_xlsx) as writer:
            for i in range(len(maps)):
                df = pd.DataFrame(tables[i])
                df.to_excel(writer, sheet_name=names[i], index=False)
            for i in range(len(im_tables)):
                df = pd.DataFrame(im_tables[i])
                df.to_excel(writer, sheet_name=im_names[i], index=False)
    return True


def data_find(data_path):
    IO_num = {}
    wb_data = openpyxl.load_workbook(data_path)
    for sheet in wb_data.sheetnames:
        ws_data = wb_data[sheet]
        count = ws_data.max_row - 2
        IO_num.update([(sheet, count)])
    return IO_num


def marker_find(sample_path, name_sheet):
    markers = []
    size = []
    start_adress = []
    number_cell_marker = -1
    number_row = 0

    wb_sample = openpyxl.load_workbook(sample_path, data_only=True)
    ws_sample = wb_sample[name_sheet]

    for row in ws_sample.rows:
        if number_cell_marker == -1:
            for cell in row:
                if cell.value == "МЕТКА":
                    number_cell_marker = cell.column
                    break
        else:
            marker_cell = row[number_cell_marker - 1]
            marker = marker_cell.value
            if marker is not None:
                markers.append(marker)
                current_row_num = row[0].row  # или row[1].row — но лучше row[0]
                size.append(current_row_num - number_row - 1)
                number_row = current_row_num
                # Получаем адрес из следующей ячейки (справа от маркера)
                if number_cell_marker < ws_sample.max_column:
                    addr_cell = row[number_cell_marker]  # столбец сразу после "МЕТКА"
                    start_adress.append(addr_cell.value)
                else:
                    start_adress.append(None)

    # === Теперь корректно обрезаем ВСЕ списки ===
    if not markers:
        return [], [], []

    # Убираем ПОСЛЕДНИЙ маркер (и соответствующие данные)
    markers = markers[:-1]
    size = size[1:]              # первый элемент size — мусорный (до первого маркера)
    start_adress = start_adress[:-1]  # ← вот чего не хватало!

    return markers, size, start_adress


def create_modbus_map(sample_map_path, out_map_path, name_sheet,
                      data_path, show_progress=False, progress_bar_tag="",
                      delete_other_sheets=False):
    try:

        IO_num = data_find(data_path)
        markers, size, start_modbus_adress = marker_find(sample_map_path, name_sheet)
        # === Проверка: все ли start_modbus_adress — числа? ===
        for idx, addr in enumerate(start_modbus_adress):
            if addr is None:
                raise ValueError(f"Начальный Modbus-адрес для маркера '{markers[idx]}' не задан (ячейка пуста или не число).")
            if not isinstance(addr, (int, float)):
                try:
                    start_modbus_adress[idx] = int(addr)
                except (ValueError, TypeError):
                    raise ValueError(f"Некорректный Modbus-адрес '{addr}' для маркера '{markers[idx]}'. Ожидалось число.")

        num_channel = [
            IO_num.get(marker.split("_")[0])
            for marker in markers
            if IO_num.get(marker.split("_")[0]) is not None
        ]

        print("IO_num:", IO_num)
        print("Markers:", markers)
        print("num_channel:", num_channel)

        if len(num_channel) != len(markers):
            raise ValueError("Несоответствие между количеством маркеров и каналами из data_path.")

        wb_data = openpyxl.load_workbook(data_path)
        wb_sample = openpyxl.load_workbook(sample_map_path)
        list_all_ws = wb_sample.sheetnames

        if delete_other_sheets:
            for item in list_all_ws:
                if item != name_sheet:
                    wb_sample.remove(wb_sample[item])

        ws_sample = wb_sample[name_sheet]
        k = 0

        for marker in markers:
            type_channel = marker.split("_")[0]
            if type_channel not in wb_data.sheetnames:
                print(f"Предупреждение: лист '{type_channel}' отсутствует в файле данных. Пропускаем маркер '{marker}'.")
                k += 1
                continue

            ws_data = wb_data[type_channel]
            row_num_data = {ch_type: 2 for ch_type in ["AI", "AO", "DI", "DO", "DVLV", "AVLV", "MTR", "BL", "MTRPID"]}
            bit_adress_modbus = 0
            row_num = 0

            for row in ws_sample.rows:
                row_num += 1
                if row[8].value == marker:
                    # === Сохраняем объединения до вставки ===
                    merged_ranges_before = list(ws_sample.merged_cells.ranges)
                    ws_sample.insert_rows(row_num + 1 + size[k], num_channel[k] * size[k])
                    ws_sample.merged_cells.ranges.clear()

                    for merged_range in merged_ranges_before:
                        min_row, min_col, max_row, max_col = (
                            merged_range.min_row,
                            merged_range.min_col,
                            merged_range.max_row,
                            merged_range.max_col,
                        )
                        if max_row < row_num + 1 + size[k]:
                            ws_sample.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
                        elif min_row < row_num + 1 + size[k] <= max_row:
                            new_max_row = max_row + num_channel[k] * size[k]
                            ws_sample.merge_cells(start_row=min_row, start_column=min_col, end_row=new_max_row, end_column=max_col)
                        elif min_row >= row_num + 1 + size[k]:
                            new_min_row = min_row + num_channel[k] * size[k]
                            new_max_row = max_row + num_channel[k] * size[k]
                            ws_sample.merge_cells(start_row=new_min_row, start_column=min_col, end_row=new_max_row, end_column=max_col)

                    # Прходимся по кол-ву сигналов конкретного типа
                    for n in range(num_channel[k] + 1):
                        #Проходимся по строчкам для одного сигнала
                        range_start = size[k] * n + row_num + 1
                        range_end = size[k] * (n + 1) + row_num + 1
                        print(range_start, range_end)
                        for j in range(range_start, range_end ):
                            ws_sample.cell(row=j, column=10).value = f"{type_channel}{n + 1}"

                            if n != num_channel[k]:
                                for col in range(11,18):
                                    old_cell = ws_sample.cell(row=j, column=col)
                                    new_cell = ws_sample.cell(row=j + size[k], column=col)
                                    is_merged = any(new_cell.coordinate in mr for mr in ws_sample.merged_cells.ranges)
                                    if not is_merged:
                                        new_cell.value = old_cell.value
                                        new_cell._style = copy(old_cell._style)

                                for i in range(1, 9):
                                    old_cell = ws_sample.cell(row=j, column=i)
                                    new_cell = ws_sample.cell(row=j + size[k], column=i)
                                    is_merged = any(new_cell.coordinate in mr for mr in ws_sample.merged_cells.ranges)
                                    if not is_merged:
                                        new_cell.value = old_cell.value
                                        new_cell._style = copy(old_cell._style)

                            for i in range(1, 9):
                                new_cell = ws_sample.cell(row=j, column=i)
                                is_merged = any(new_cell.coordinate in mr for mr in ws_sample.merged_cells.ranges)
                                if not is_merged:
                                    if i == 1:
                                        old_value = str(new_cell.value)
                                        desc1 = ws_data.cell(row=row_num_data[type_channel], column=1).value or ""
                                        desc2 = ws_data.cell(row=row_num_data[type_channel], column=2).value or ""
                                        replacement = f"{desc1}-{desc2}"
                                        new_value = old_value.replace("$$", replacement)
                                        new_cell.value = new_value
                                    elif i == 2:
                                        # === Безопасное обновление адреса ===
                                        current_addr = start_modbus_adress[k]
                                        if current_addr is None:
                                            raise RuntimeError(f"Адрес стал None для маркера {marker} на шаге k={k}")

                                        cell_data_type = str(ws_sample.cell(row=j, column=4).value or "").lower()
                                        next_cell_type = ""
                                        if j + 1 <= ws_sample.max_row:
                                            next_cell_type = str(ws_sample.cell(row=j + 1, column=4).value or "").lower()

                                        if "4 byte" in cell_data_type:
                                            size_val_modbus = 2
                                            bit_modbus = 0
                                            if "bool" in next_cell_type:
                                                start_modbus_adress[k] += size_val_modbus
                                            ws_sample.cell(row=j, column=2).value = str(current_addr)
                                        elif "2 byte" in cell_data_type:
                                            size_val_modbus = 1
                                            bit_modbus = 0
                                            if "bool" in next_cell_type:
                                                start_modbus_adress[k] += size_val_modbus
                                            ws_sample.cell(row=j, column=2).value = str(current_addr)
                                        elif "bool" in cell_data_type:
                                            size_val_modbus = 0
                                            bit_modbus = 1
                                            if "4 byte" in next_cell_type:
                                                size_val_modbus = -1
                                            ws_sample.cell(row=j, column=2).value = f"{current_addr}.{bit_adress_modbus}"
                                        else:
                                            ws_sample.cell(row=j, column=2).value = f"{current_addr}.{bit_adress_modbus}"

                                    elif i == 6:
                                        name_val = ws_data.cell(row=row_num_data[type_channel], column=5).value or ""
                                        unit_val = ws_data.cell(row=row_num_data[type_channel], column=4).value or ""
                                        channel_name = f"{name_val}({unit_val})"
                                        ws_sample.cell(row=j, column=i).value = channel_name

                            # === Безопасное обновление адресов ===
                            try:
                                start_modbus_adress[k] += size_val_modbus
                            except Exception as e:
                                raise RuntimeError(f"Ошибка при обновлении адреса для маркера {marker}, k={k}: {e}")

                            bit_adress_modbus += bit_modbus

                            # Условия сброса битового адреса
                            if ((size[k] > 1 and bit_adress_modbus >= (4 * size[k])) or
                                bit_adress_modbus > 15 or
                                (type_channel == "DVLV" and bit_adress_modbus >= size[k]) or
                                (type_channel == "AI" and bit_adress_modbus >= size[k]) or
                                (size[k] > 0 and 16 // size[k] == 1 and bit_adress_modbus >= size[k])):
                                start_modbus_adress[k] += 1
                                bit_adress_modbus = 0

                        row_num_data[type_channel] += 1

                        if row_num_data[type_channel] > ws_data.max_row:
                            print(
                                f"⚠️ Попытка чтения за пределами данных: строка {row_num_data[type_channel]}, максимум {ws_data.max_row}")
                            break

            k += 1
            if show_progress:
                progress = k / len(markers)
                dpg.set_value(progress_bar_tag, progress)
                dpg.configure_item(progress_bar_tag, show=True, overlay=f"{round(progress * 100, 1)}%")
                print(f"Обработан маркер: {marker} ({k}/{len(markers)})")

                if progress == 1:
                    dpg.configure_item(progress_bar_tag, show=False)

        wb_sample.save(out_map_path)
        return True

    except Exception as e:
        print(f"❌ Ошибка в create_modbus_map: {e}")
        raise  # или return False, если не хотите прерывать


# ============ GUI ============

global path_data
global path_sample
global path_export
global paths
paths = ["", "", ""]
global sample_list
global sample_name
global data_numbers
data_numbers = {"start": 1, "end": 1}
global error_str
error_str = ""
global sample_sheetnames
sample_sheetnames = []
global generate_address_labels, generate_name_table, generate_journal, generate_sample, generate_macros_text
generate_address_labels = False
generate_name_table = False
generate_journal = False
generate_sample = False
generate_macros_text = False

dpg.create_context()

# Зелёный для успеха
with dpg.theme() as success_theme:
    with dpg.theme_component(dpg.mvText):
        dpg.add_theme_color(dpg.mvThemeCol_Text, (0, 255, 0), category=dpg.mvThemeCat_Core)  # RGB: Зелёный

# Красный для ошибок/предупреждений
with dpg.theme() as error_theme:
    with dpg.theme_component(dpg.mvText):
        dpg.add_theme_color(dpg.mvThemeCol_Text, (255, 0, 0), category=dpg.mvThemeCat_Core)  # RGB: Красный

def handle_overwrite_response(filename, on_confirm, on_cancel):
    """Показывает диалог перезаписи и вызывает колбэк после выбора."""
    def on_yes():
        dpg.delete_item("overwrite_modal")
        on_confirm(filename)

    def on_no():
        dpg.delete_item("overwrite_modal")
        on_cancel()

    def on_rename(sender, app_data):
        nonlocal filename
        new_name = app_data.strip()
        if new_name:
            filename = os.path.join(os.path.dirname(filename), new_name)

    with dpg.mutex():
        with dpg.window(label="Файл существует", modal=True, no_close=True, tag="overwrite_modal"):
            dpg.add_text(f"Файл {os.path.basename(filename)} уже существует.")
            dpg.add_text("Перезаписать?")
            dpg.add_input_text(default_value=os.path.basename(filename), callback=on_rename, width=300, tag="rename_input")
            with dpg.group(horizontal=True):
                dpg.add_button(label="Да", callback=on_yes)
                dpg.add_button(label="Нет", callback=on_no)

def create_map(sender, appdata):
    global sample_name
    dpg.configure_item("load_indic_2", show=True)
    try:
        if not paths[1]:
            raise ValueError("Не выбран файл шаблона карты Modbus")
        if not sample_name:
            raise ValueError("Не выбран лист шаблона")
        if sample_name not in sample_sheetnames:
            raise ValueError(f"Выбранный лист '{sample_name}' отсутствует в файле шаблона. Перезагрузите шаблон.")

        if not paths[2]:
            raise ValueError("Не указан путь для сохранения")

        signals_path = os.path.join(paths[2], "signals.xlsx")
        if not os.path.exists(signals_path):
            raise FileNotFoundError("Файл signals.xlsx не найден. Сначала сформируйте список сигналов.")

        out_map_path = os.path.join(paths[2], "modbus_map.xlsx")

        def continue_with_path(final_path):
            try:
                create_modbus_map(paths[1], final_path, sample_name, signals_path,
                                  show_progress=True, progress_bar_tag="progress_bar_map")
                set_colored_message([("Карта сохранена: " + final_path, success_theme)])
            except Exception as ex:
                set_colored_message([(str(ex), error_theme)])
            finally:
                dpg.configure_item("load_indic_2", show=False)

        def on_cancel():
            set_colored_message([("Операция отменена пользователем.", error_theme)])
            dpg.configure_item("load_indic_2", show=False)

        if os.path.exists(out_map_path):
            handle_overwrite_response(out_map_path, continue_with_path, on_cancel)
        else:
            continue_with_path(out_map_path)

    except Exception as ex:
        set_colored_message([(str(ex), error_theme)])
        dpg.configure_item("load_indic_2", show=False)

def create_xlsx(sender, app_data):
    dpg.configure_item("load_indic_1", show=True)
    try:
        if not paths[0]:
            raise ValueError("Не выбран файл КД")
        if not paths[0].lower().endswith('.pdf'):
            raise ValueError("Файл КД должен быть в формате PDF")
        if not paths[2]:
            raise ValueError("Не указан путь для сохранения")

        signals_path = os.path.join(paths[2], "signals.xlsx")

        # Функция продолжения после выбора
        def continue_with_path(final_path):
            try:
                pdf_to_xlsx(paths[0], final_path, data_numbers["start"], data_numbers["end"])
                set_colored_message([("Сигналы сохранены: " + final_path, success_theme)])
            except Exception as ex:
                set_colored_message([(str(ex), error_theme)])
            finally:
                dpg.configure_item("load_indic_1", show=False)

        def on_cancel():
            set_colored_message([("Операция отменена пользователем.", error_theme)])
            dpg.configure_item("load_indic_1", show=False)

        if os.path.exists(signals_path):
            handle_overwrite_response(signals_path, continue_with_path, on_cancel)
        else:
            continue_with_path(signals_path)

    except Exception as ex:
        set_colored_message([(str(ex), error_theme)])
        dpg.configure_item("load_indic_1", show=False)

def create_table(sender, appdata):
    try:
        if not paths[2]:
            raise ValueError("Не указан путь для сохранения")

        # --- Логика выбора файла modbus_for_panel.xlsx ---
        modbus_panel_path = None
        if generate_journal: # Проверяем, нужно ли генерировать журнал
            # Проверяем, был ли выбран чекбокс "Адресные метки"
            address_labels_selected = dpg.get_value("cb_address") # Получаем текущее состояние чекбокса
            if address_labels_selected and generate_address_labels: # Если чекбокс был установлен и его состояние True
                # Используем файл, созданный convert_modbus_map
                modbus_panel_path = os.path.join(paths[2], "modbus_for_panel.xlsx")
                # Проверим, существует ли файл, на всякий случай
                if not os.path.exists(modbus_panel_path):
                     # Если файл не найден, всё равно запрашиваем у пользователя
                     modbus_panel_path = select_file_system_dialog(
                         filetypes=[("Excel файлы", "*.xlsx")],
                         initial_dir=paths[2],
                         title="Файл 'modbus_for_panel.xlsx' не найден. Выберите файл карты регистров панели."
                     )
            else:
                # Запрашиваем у пользователя файл
                modbus_panel_path = select_file_system_dialog(
                    filetypes=[("Excel файлы", "*.xlsx")],
                    initial_dir=paths[2],
                    title="Выберите файл карты регистров панели (modbus_for_panel.xlsx)."
                )

            if not modbus_panel_path or not os.path.exists(modbus_panel_path):
                if generate_journal: # Если пользователь отменил и мы действительно хотим журнал
                    raise FileNotFoundError("Файл карты регистров панели не выбран или не найден. Невозможно сгенерировать журнал.")

        # --- Логика выбора файла signals.xlsx ---
        signals_path = None
        if (generate_sample or generate_name_table or generate_macros_text):
            signals_path = select_file_system_dialog(
                filetypes=[("Excel файлы", "*.xlsx")],
                initial_dir=paths[2],
                title="Выберите ваш файл с таблицей сигналов (signals.xlsx)"
            )
            if not signals_path or not os.path.exists(signals_path):
                 # Проверим, нужно ли *обязательно* signals.xlsx для генерации журнала.
                 # generate_alarms_from_modbus_map использует modbus_for_panel.xlsx.
                 # Если для журнала нужен signals.xlsx, добавь проверку здесь.
                 # Если нет, и журнал может работать только с modbus_panel_path, можно оставить как есть для других опций.
                 # Предположим, что signals.xlsx НЕ нужен для generate_alarms_from_modbus_map, но нужен для других.
                 # Проверим, нужен ли он для других опций.
                 if (generate_sample or generate_name_table or generate_macros_text) and not signals_path:
                    raise FileNotFoundError("Файл c signals.xlsx не найден. Необходим для других опций.")
                 # Если generate_journal = True, но signals не нужен для него, просто продолжим
                 # Но если signals всё-таки нужен для журнала, раскомментируй следующие строки:
                 # if generate_journal and not signals_path:
                 #    raise FileNotFoundError("Файл c signals.xlsx не найден. Необходим для генерации журнала.")


        # --- Логика генерации ---
        messages = []
        if not (generate_address_labels or generate_name_table
                or generate_journal or generate_sample
                or generate_macros_text):
            raise ValueError("Не выбраны таблицы для генерации")

        # 1. Таблица строк
        if generate_name_table:
            if not signals_path: # Проверяем, что signals_path доступен, если нужно
                 raise ValueError("Файл signals.xlsx не выбран, но необходим для генерации Таблицы строк.")
            try:
                create_name_table(
                    out_map_path=paths[2],
                    data_path=signals_path,
                    file_name="name_table.xlsx",
                    name_order=list(range(0, 50)),
                    start_id=0,
                    enable_export=True
                )
                messages.append("Таблица строк создана")
            except PermissionError:
                raise PermissionError("Файл name_table.xlsx уже открыт в Excel. Закройте его и повторите попытку.")

        # 2. Адресные метки
        if generate_address_labels:
             # signals_path не нужен для convert_modbus_map, modbus_panel_path не используется здесь
             # но файл signals.xlsx нужен для create_modbus_map (ранее), который генерирует modbus_map.xlsx
             # convert_modbus_map берет modbus_map.xlsx и создает modbus_for_panel.xlsx
             # Предполагаем, что modbus_map.xlsx уже создан.
             # Путь к modbus_map.xlsx (результат create_modbus_map)

             # Запрашиваем у пользователя файл
             modbus_map_path = select_file_system_dialog(
                 filetypes=[("Excel файлы", "*.xlsx")],
                 initial_dir=paths[2],
                 title="Выберите файл карты регистров  (modbus_map.xlsx)."
             )

             #modbus_map_path = os.path.join(paths[2], "modbus_map.xlsx")

             if not os.path.exists(modbus_map_path):
                 raise FileNotFoundError("Файл modbus_map.xlsx не найден. Сначала сформируйте карту регистров.")
             try:
                success = convert_modbus_map(
                    path_map=modbus_map_path,
                    name_new_map=os.path.join(paths[2], "modbus_for_panel"), # Указываем путь без расширения
                    name_sheet="Шаблон" # Или передавать имя листа, если оно может меняться
                )
                if success:
                    messages.append("Адресные метки созданы")
                else:
                    raise RuntimeError("Ошибка при создании адресных меток")
             except PermissionError:
                raise PermissionError("Файл modbus_for_panel.xlsx уже открыт в Excel. Закройте его и повторите попытку.")

        # 3. Журнал тревог
        if generate_journal:
            # modbus_panel_path уже определен выше
            # signals_path может быть не нужен, проверь generate_alarms_from_modbus_map
            output_alarm_file = os.path.join(paths[2], "journal_for_panel.xlsx") # Имя выходного файла
            try:
                # Вызываем функцию из другого файла
                # Предполагаем, что generate_alarms_from_modbus_map принимает путь к карте и путь к выходному файлу
                generate_alarms_from_modbus_map(modbus_panel_path, output_alarm_file)
                messages.append("Журнал тревог создан")
            except PermissionError:
                messages.append(
                    "Файл journal_for_panel.xlsx уже открыт в Excel. Закройте его и повторите попытку."
                )
            except Exception as e:
                # Выводим более подробную ошибку
                messages.append(f"Ошибка при создании журнала тревог: {str(e)}")


        # 4. Таблица выборки
        if generate_sample:
            if not signals_path: # Проверяем, что signals_path доступен
                 raise ValueError("Файл signals.xlsx не выбран, но необходим для генерации Таблицы выборки.")
            try:
                success = generate_sampling_table(
                    input_path=signals_path,
                    output_path=os.path.join(paths[2], "sample_for_panel.xlsx")
                )
                if success:
                    messages.append("Таблица выборки создана")
                else:
                    messages.append("Функция generate_sampling_table вернула False")
            except PermissionError:
                messages.append(
                    "Файл sample_for_panel.xlsx уже открыт в Excel. Закройте его и повторите попытку."
                )
            except Exception as e:
                messages.append(f"Ошибка при создании таблицы выборки: {str(e)}")

        # 5. Файлы макросов
        if generate_macros_text:
            if not signals_path: # Проверяем, что signals_path доступен
                 raise ValueError("Файл signals.xlsx не выбран, но необходим для генерации файлов макросов.")
            try:
                success = generate_all_case_files(
                    input_excel=signals_path,
                    output_dir=os.path.join(paths[2], "macros_text"),
                    max_len=40
                )
                if success:
                    messages.append("Файлы для макросов созданы")
                else:
                    messages.append("Функция generate_all_case_files вернула False")
            except PermissionError:
                messages.append(
                    "Файл или папка макросов заняты. Закройте Excel или проводник и повторите попытку."
                )
            except Exception as e:
                messages.append(f"Ошибка при создании файлов макросов: {str(e)}")

        if messages:
            colored_messages = []
            for msg in messages:
                if "создан" in msg.lower() or "созданы" in msg.lower():  # Проверяем на успешные слова
                    colored_messages.append((msg + "; ", success_theme))
                else:  # Остальные (ошибки, предупреждения, "Функция вернула False")
                    colored_messages.append((msg + "; ", error_theme))
            set_colored_message(colored_messages)
        else:
            # Для сообщения "Ничего не сгенерировано" можно выбрать цвет, например, красный как предупреждение
            set_colored_message([("Ничего не сгенерировано", error_theme)])

    except Exception as ex:
        set_colored_message([(f"Ошибка: {str(ex)}", error_theme)])

def set_numbers(sender, app_data):
    data_numbers[sender] = app_data

def set_sample_name(sender, app_data):
    global sample_name
    sample_name = str(app_data)

def show_sheet_selection_dialog(sheetnames):
    """Показывает модальное окно для выбора листа из списка."""
    def on_select(sender, app_data):
        global sample_name
        selected = dpg.get_value("sheet_selector_combo")
        if selected:
            sample_name = selected
            dpg.configure_item("sample_combo"
                               , items=sheetnames)
            dpg.set_value("sample_combo", selected)
            set_colored_message([("Выбран лист: " + selected, success_theme)])
        dpg.delete_item("sheet_selection_modal")

    def on_cancel():
        dpg.delete_item("sheet_selection_modal")
        set_colored_message([("Не выбран лист шаблона. Выберите файл снова.", error_theme)])

    with dpg.mutex():
        with dpg.window(label="Выберите лист шаблона", modal=True, no_close=True, tag="sheet_selection_modal"):
            dpg.add_text("Лист 'Шаблон' не найден. Выберите нужный лист:")
            dpg.add_combo(items=sheetnames, default_value=sheetnames[0] if sheetnames else "", tag="sheet_selector_combo", width=250)
            with dpg.group(horizontal=True):
                dpg.add_button(label="Выбрать", callback=on_select)
                dpg.add_button(label="Отмена", callback=on_cancel)

def sample_list_extr(sender, app_data):
    global sample_list, sample_name, sample_sheetnames
    try:
        path_sample = paths[1]
        if not path_sample.lower().endswith('.xlsx'):
            raise ValueError("Файл шаблона должен быть в формате XLSX")
        wb_sample = openpyxl.load_workbook(path_sample, read_only=True)
        sheetnames = wb_sample.sheetnames
        wb_sample.close()
        sample_sheetnames = sheetnames  # Сохраняем актуальный список

        if "Шаблон" in sheetnames:
            sample_name = "Шаблон"
            dpg.configure_item("sample_combo", items=sheetnames)
            dpg.set_value("sample_combo", "Шаблон")
            set_colored_message([("Автоматически выбран лист: Шаблон", success_theme)])
        else:
            # Показываем диалог выбора листа
            show_sheet_selection_dialog(sheetnames)
    except Exception as ex:
        set_colored_message([(str(ex), error_theme)])

def get_pdf_page_count(pdf_path):
    try:
        with pl.open(pdf_path) as pdf:
            return len(pdf.pages)
    except Exception as ex:
        print(f"Ошибка при чтении PDF: {ex}")
        return 1

def path_extractor(sender, app_data, id):
    global sample_name, sample_sheetnames
    try:
        if sender == "filedialog_2":
            paths[id] = app_data["current_path"]
        else:
            paths[id] = str(list(app_data["selections"].values())[0])
        dpg.set_value(f"input_text_{id}", paths[id])

        # === Обработка PDF (id=0) ===
        if id == 0:
            if not paths[0].lower().endswith('.pdf'):
                set_colored_message([("Файл КД должен иметь расширение .pdf", error_theme)])
                paths[0] = ""  # сбрасываем путь
                dpg.set_value("input_text_0", "")
                return
            # Только если это PDF — получаем количество страниц
            total_pages = get_pdf_page_count(paths[0])
            dpg.configure_item("start", max_value=total_pages)
            dpg.configure_item("end", max_value=total_pages)
            dpg.set_value("end", total_pages)
            data_numbers["end"] = total_pages

        # === Обработка XLSX шаблона (id=1) ===
        elif id == 1:
            sample_name = ""
            sample_sheetnames = []
            if not paths[1].lower().endswith('.xlsx'):
                set_colored_message([("Файл шаблона должен иметь расширение .xlsx")])
                paths[1] = ""  # сбрасываем путь
                dpg.set_value("input_text_1", "")
                return
            # Автоматически загружаем листы
            sample_list_extr(None, None)

        if id == 2:  # путь сохранения
            dpg.configure_item("btn_generate_tables", enabled=bool(paths[2]))

    except Exception as ex:
        set_colored_message([(str(ex), error_theme)])

def select_file(name, id, only_directory=False):
    global sample_name, sample_sheetnames
    try:
        # Определяем начальную директорию
        initial_dir = None
        if paths[id]:
            if only_directory:
                initial_dir = paths[id]
            else:
                # Если это файл, то извлекаем директорию
                if os.path.isfile(paths[id]):
                    initial_dir = os.path.dirname(paths[id])
                else:
                    initial_dir = paths[id]

        if only_directory:
            path = select_folder_system_dialog(initial_dir=initial_dir)
        else:
            if id == 0:  # файл КД (PDF)
                filetypes = [("PDF файлы", "*.pdf")]
            elif id == 1:  # файл шаблона (XLSX)
                filetypes = [("Excel файлы", "*.xlsx")]
            else:
                filetypes = [("Все файлы", "*.*")]
            path = select_file_system_dialog(
                filetypes=filetypes,
                initial_dir=initial_dir
            )

        if not path:
            return  # пользователь отменил

        # === Обработка PDF (id=0) ===
        if id == 0:
            if not path.lower().endswith('.pdf'):
                set_colored_message([("Файл КД должен иметь расширение .pdf", error_theme)])
                return
            # Только если это PDF — получаем количество страниц
            total_pages = get_pdf_page_count(path)
            dpg.configure_item("start", max_value=total_pages)
            dpg.configure_item("end", max_value=total_pages)
            dpg.set_value("end", total_pages)
            data_numbers["end"] = total_pages

        # === Обработка XLSX шаблона (id=1) ===
        elif id == 1:
            sample_name = ""
            sample_sheetnames = []
            if not path.lower().endswith('.xlsx'):
                set_colored_message([("Файл шаблона должен иметь расширение .xlsx", error_theme)])
                return
            # Автоматически загружаем листы
            paths[id] = path
            dpg.set_value(f"input_text_{id}", paths[id])
            sample_list_extr(None, None)
            return  # чтобы не сбрасывался путь ниже

        # === Сохраняем путь и обновляем интерфейс ===
        paths[id] = path
        dpg.set_value(f"input_text_{id}", paths[id])

        # === Обновляем кнопку генерации таблиц ===
        if id == 2:  # путь сохранения
            dpg.configure_item("btn_generate_tables", enabled=bool(paths[2]))

    except Exception as ex:
        set_colored_message([(str(ex), error_theme)])

def on_address_labels_checkbox(sender, app_data):
    if not paths[2]:
        dpg.set_value(sender, False)  # сбрасываем чекбокс
        set_colored_message([("Сначала укажите путь для сохранения", error_theme)])
        return
    global generate_address_labels
    generate_address_labels = app_data

def on_name_table_checkbox(sender, app_data):
    if not paths[2]:
        dpg.set_value(sender, False)
        set_colored_message([("Сначала укажите путь для сохранения", error_theme)])
        return
    global generate_name_table
    generate_name_table = app_data

def on_journal_checkbox(sender, app_data):
    if not paths[2]:
        dpg.set_value(sender, False)
        set_colored_message([("Сначала укажите путь для сохранения", error_theme)])
        return
    global generate_journal
    generate_journal = app_data

def on_journal_checkbox(sender, app_data):
    if not paths[2]:
        dpg.set_value(sender, False)
        set_colored_message([("Сначала укажите путь для сохранения", error_theme)])
        return

    global generate_journal
    generate_journal = app_data

def on_sample_table(sender, app_data):
    if not paths[2]:
        dpg.set_value(sender, False)
        set_colored_message([("Сначала укажите путь для сохранения", error_theme)])
        return
    global generate_sample
    generate_sample = app_data

def on_macros_text(sender, app_data):
    if not paths[2]:
        dpg.set_value(sender, False)
        set_colored_message([("Сначала укажите путь для сохранения", error_theme)])
        return
    global generate_macros_text
    generate_macros_text = app_data

def show_sheet_selection_dialog_for_generation(sheetnames, on_confirm):
    """Показывает диалог выбора листа и вызывает on_confirm с выбранным именем."""
    selected_sheet = [None]

    def on_select():
        selected = dpg.get_value("sheet_selector_combo_gen")
        if selected:
            selected_sheet[0] = selected
        dpg.delete_item("sheet_selection_modal_gen")
        if selected:
            on_confirm(selected)

    def on_cancel():
        dpg.delete_item("sheet_selection_modal_gen")

    with dpg.mutex():
        with dpg.window(label="Выберите лист из modbus_map.xlsx", modal=True, no_close=True, tag="sheet_selection_modal_gen"):
            dpg.add_text("Выберите лист ШАБЛОН, на основе которого генерировать таблицы:")
            dpg.add_combo(items=sheetnames, default_value=sheetnames[0] if sheetnames else "", tag="sheet_selector_combo_gen", width=250)
            with dpg.group(horizontal=True):
                dpg.add_button(label="ОК", callback=on_select)
                dpg.add_button(label="Отмена", callback=on_cancel)

def select_file_system_dialog(filetypes=None, initial_dir=None, title="Выберите файл"):
    """Открывает системный диалог для выбора файла."""
    root = tk.Tk()
    root.withdraw()  # скрываем главное окно
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes or [("Все файлы", "*.*")],
        initialdir=initial_dir
    )
    root.destroy()
    return file_path

def select_folder_system_dialog(initial_dir=None, title="Выберите папку"):
    """Открывает системный диалог для выбора папки."""
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(
        title=title,
        initialdir=initial_dir
    )
    root.destroy()
    return folder_path

def make_callback(name, id, only_directory):
    def callback(s, a):
        select_file(name, id, only_directory)
    return callback

def set_colored_message(messages_list):
    """
    Очищает контейнер сообщений и добавляет туда список кортежей (сообщение, цвет),
    где цвет - это тег темы (например, success_theme или error_theme).
    """
    # Удаляем все дочерние элементы из контейнера сообщений
    for item in dpg.get_item_children("message_container", 1):
        dpg.delete_item(item)

    # Создаём горизонтальную группу для размещения фрагментов текста
    with dpg.group(horizontal=True, parent="message_container", tag="message_group"):
        for msg, theme_tag in messages_list:
            # Создаём текстовый элемент
            text_tag = f"text_part_{len(dpg.get_item_children('message_group', 1))}" # Уникальный тег
            with dpg.group(horizontal=False):
                wrap_width = 600
                dpg.add_text(default_value=msg, tag=text_tag, wrap=wrap_width)
            # Привязываем тему к элементу

            print(text_tag, theme_tag)
            dpg.bind_item_theme(text_tag, theme_tag)


def gui():
    window_width = 800
    window_height = 800


    dpg.create_viewport(title='Modbus App', width=window_width, height=window_height, resizable=False)

    with dpg.font_registry():
        with dpg.font('C:\\Windows\\Fonts\\arial.ttf', 18, default_font=True, id="Default font"):
            dpg.add_font_range_hint(dpg.mvFontRangeHint_Cyrillic)
    dpg.bind_font("Default font")

    with dpg.window(label="Главное окно", pos=(0, 0), width=window_width, height=window_height, no_resize=True,
                    no_move=True, no_collapse=True, no_close=True, no_title_bar=True, no_bring_to_front_on_focus=True):

        with dpg.collapsing_header(label="1. Источник данных(КД)", leaf=True):
            with dpg.group(horizontal=True):
                dpg.add_text("Выберите файл с конструкторской документацией(КД):")
                dpg.add_text("(PDF)", color=(255, 0, 0))
            with dpg.group(horizontal=True):
                dpg.add_button(label="Выбрать",
                               callback=make_callback( "Путь к файлу с данными", 0, False))
                dpg.add_input_text(hint="Путь к файлу с данными", tag="input_text_0", width=690)

            dpg.add_text("Выберите диапозон страниц ТОЛЬКО с таблицей сигналов в КД:")
            with dpg.group(horizontal=True):
                dpg.add_text("Начало с ")
                dpg.add_input_int(tag="start", width=100, callback=set_numbers, min_clamped=True, min_value=1,
                                  max_clamped=True, max_value=300, default_value=1)
                dpg.add_text(" до ")
                dpg.add_input_int(tag="end", width=100, callback=set_numbers, min_clamped=True, min_value=1,
                                  max_clamped=True, max_value=300, default_value=1)
                dpg.add_text("страницы(включительно).")

        with dpg.collapsing_header(label="2. Шаблон карты Modbus", leaf=True):
            with dpg.group(horizontal=True):
                dpg.add_text("Выберите файл XLSX с шаблонами карт:")
                dpg.add_text("(XLSX)", color=(0, 255, 0))
            with dpg.group(horizontal=True):
                dpg.add_button(label="Выбрать",
                               callback=make_callback("Путь к файлу с шаблоном", 1, False))
                dpg.add_input_text(hint="Путь к файлу с шаблоном", tag="input_text_1", width=690)
            with dpg.group(horizontal=True):
                dpg.add_text("Выберите шаблон из списка:")
                dpg.add_combo(width=200, callback=set_sample_name, tag="sample_combo")

        with dpg.collapsing_header(label="3. Обработка и путь сохранения", leaf=True):
            dpg.add_text("В данную папку будет сохранен файл со списком сигналов(signals.xlsx)"
                         + " и файл карты регистров(modbus_map.xlsx).", wrap=550)
            with dpg.group(horizontal=True):
                dpg.add_button(label="Выбрать",
                               callback=make_callback("Путь к папке с проектом", 2, True))
                dpg.add_input_text(hint="Путь к папке с проектом", tag="input_text_2", width=690)

        with dpg.collapsing_header(label="4. Данные для визуализации", leaf=True):
            dpg.add_text("Выберите необходимые файлы для визуализации Weintek EasyBuilderPro")
            with dpg.group(horizontal=True):
                dpg.add_checkbox(label="Адресные метки", callback=on_address_labels_checkbox, tag="cb_address")
                dpg.add_checkbox(label="Таблица строк", callback=on_name_table_checkbox, tag="cb_name_table")
                dpg.add_checkbox(label="Журнал тревог", callback=on_journal_checkbox, tag="cb_journal")
                dpg.add_checkbox(label="Таблица выборки", callback=on_sample_table, tag="cb_sample_table")
                dpg.add_checkbox(label="Файлы макросв", callback=on_macros_text, tag="cb_macros_text")
            with dpg.group(horizontal=True):
                dpg.add_button(label="Сформировать файлы", callback=create_table, tag="btn_generate_tables")
                dpg.configure_item("btn_generate_tables", enabled=bool(paths[2]))

        with dpg.group(horizontal=True):
            dpg.add_button(label="Сформировать список сигналов", callback=create_xlsx, pos=(8, window_height - 160),
                           width=250)
            dpg.add_loading_indicator(style=1, radius=1.3, color=(0, 0, 255), show=False, tag="load_indic_1")

        with dpg.group(horizontal=True):
            dpg.add_button(label="Сформировать карту регистров", callback=create_map, pos=(8, window_height - 130),
                           width=250)
            dpg.add_loading_indicator(style=1, radius=1.3, color=(0, 0, 255), show=False, tag="load_indic_2")
            dpg.add_progress_bar(tag="progress_bar_map", show=False, width=200)

        with dpg.child_window(tag="message_container", pos=(8, window_height - 90), width=800 - 32, height=60,
                              horizontal_scrollbar=True):
            dpg.add_text(default_value="", tag="error_text_dynamic")

    with dpg.theme() as global_theme:
        with dpg.theme_component(dpg.mvAll):
            dpg.add_theme_color(dpg.mvThemeCol_FrameBg, (130, 112, 90), category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 5, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_TextDisabled, (255, 255, 255), category=dpg.mvThemeCat_Core)

        with dpg.theme_component(dpg.mvInputInt):
            dpg.add_theme_color(dpg.mvThemeCol_FrameBg, (114, 130, 90), category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 5, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_Text, (255, 255, 255), category=dpg.mvThemeCat_Core)

        with dpg.theme_component(dpg.mvButton):
            dpg.add_theme_color(dpg.mvThemeCol_Button, (143, 143, 143), category=dpg.mvThemeCat_Core)

    with dpg.theme() as item_theme:
        with dpg.theme_component(dpg.mvInputText):
            dpg.add_theme_color(dpg.mvThemeCol_FrameBg, (69, 69, 69), category=dpg.mvThemeCat_Core)
            dpg.add_theme_style(dpg.mvStyleVar_FrameRounding, 5, category=dpg.mvThemeCat_Core)
            dpg.add_theme_color(dpg.mvThemeCol_Text, (196, 135, 135), category=dpg.mvThemeCat_Core)



    dpg.bind_theme(global_theme)

    dpg.setup_dearpygui()
    dpg.show_viewport()
    dpg.start_dearpygui()
    dpg.destroy_context()


if __name__ == '__main__':
    gui()